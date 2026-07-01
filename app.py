from flask import Flask, render_template, request, redirect, session, url_for, flash, send_file, send_from_directory
from flask_mail import Mail, Message
import os
import json
import urllib.parse
from collections import Counter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.utils import secure_filename
from flask_sqlalchemy import SQLAlchemy
import pandas as pd
from sqlalchemy import text
from werkzeug.security import generate_password_hash, check_password_hash
from io import BytesIO
import unicodedata
from datetime import datetime, date




app = Flask(__name__)

app.config["SECRET_KEY"] = os.getenv(
    "SECRET_KEY",
    "dev_secret_key"
)
app.config["MAIL_SERVER"] = os.getenv("MAIL_SERVER")
app.config["MAIL_PORT"] = int(os.getenv("MAIL_PORT"))
app.config["MAIL_USE_TLS"] = os.getenv("MAIL_USE_TLS").lower() == "true"
app.config["MAIL_USERNAME"] = os.getenv("MAIL_USERNAME")
app.config["MAIL_PASSWORD"] = os.getenv("MAIL_PASSWORD")
app.config["MAIL_DEFAULT_SENDER"] = os.getenv("MAIL_DEFAULT_SENDER")

mail = Mail(app)

@app.context_processor
def carregar_notificacoes():

    usuario = session.get("usuario")

    if not usuario:
        return {}

    notificacoes = db.session.execute(
        text("""
            SELECT
                id,
                titulo,
                mensagem,
                tipo,
                link,
                lida,
                criada_em
            FROM notificacoes
            WHERE usuario_id = :usuario_id
            ORDER BY criada_em DESC
            LIMIT 5
        """),
        {
            "usuario_id": usuario["id"]
        }
    ).fetchall()

    total_notificacoes = db.session.execute(
        text("""
            SELECT COUNT(*)
            FROM notificacoes
            WHERE usuario_id = :usuario_id
            AND lida = false
        """),
        {
            "usuario_id": usuario["id"]
        }
    ).scalar()

    return {
        "notificacoes_topo": notificacoes,
        "total_notificacoes": total_notificacoes
    }


def criar_notificacao(
    usuario_id,
    titulo,
    mensagem,
    tipo="sistema",
    link="#"
):

    db.session.execute(
        text("""
            INSERT INTO notificacoes
            (
                usuario_id,
                titulo,
                mensagem,
                tipo,
                link
            )
            VALUES
            (
                :usuario_id,
                :titulo,
                :mensagem,
                :tipo,
                :link
            )
        """),
        {
            "usuario_id": usuario_id,
            "titulo": titulo,
            "mensagem": mensagem,
            "tipo": tipo,
            "link": link
        }
    )

    db.session.commit()


@app.route("/notificacao/<int:id>")
def abrir_notificacao(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    db.session.execute(
        text("""
            UPDATE notificacoes
            SET lida = true
            WHERE id = :id
            AND usuario_id = :usuario_id
        """),
        {
            "id": id,
            "usuario_id": usuario["id"]
        }
    )

    db.session.commit()

    return redirect(request.args.get("link", "/"))

# =========================
# EXCLUIR NOTIFICAÇÃO
# =========================

@app.route("/excluir_notificacao/<int:id>")
def excluir_notificacao(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    db.session.execute(

        db.text("""

            DELETE FROM notificacoes

            WHERE id = :id
            AND usuario_id = :usuario_id

        """),

        {
            "id": id,
            "usuario_id": usuario["id"]
        }

    )

    db.session.commit()

    return redirect(request.referrer or "/")


# =========================
# LIMPAR NOTIFICAÇÕES
# =========================

@app.route("/limpar_notificacoes")
def limpar_notificacoes():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    db.session.execute(

        db.text("""

            DELETE FROM notificacoes

            WHERE usuario_id = :usuario_id

        """),

        {
            "usuario_id": usuario["id"]
        }

    )

    db.session.commit()

    return redirect(request.referrer or "/")

# =========================
# ENVIO EMAIL
# =========================
def enviar_email(destinatario, assunto, html):

    try:

        msg = Message(
            subject=assunto,
            recipients=[destinatario]
        )

        msg.html = html

        mail.send(msg)

        print(f"Email enviado para {destinatario}")

        return True

    except Exception as e:

        print("ERRO AO ENVIAR EMAIL:")
        print(e)

        return False

# =========================
# CONFIG POSTGRESQL
# =========================

app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
app.secret_key = os.getenv("SECRET_KEY")
UPLOAD_FOLDER = "uploads"

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route("/uploads/<path:filename>")
def uploaded_file(filename):

    return send_from_directory(
        app.config["UPLOAD_FOLDER"],
        filename
    )
# =========================
# LOGIN
# =========================

@app.route(
    "/login",
    methods=["GET", "POST"]
)
def login():

    if request.method == "POST":

        nome = request.form.get("nome")

        senha = request.form.get("senha")

        resultado = db.session.execute(

            db.text("""

                SELECT *

                FROM usuarios

                WHERE nome = :nome
                AND ativo = true

            """),

            {
                "nome": nome
            }

        ).fetchone()

        if resultado and check_password_hash(resultado.senha, senha):

            session["usuario"] = {

                "id": resultado.id,

                "nome": resultado.nome,

                "tipo": resultado.tipo,

                "setor": resultado.setor

            }
            if resultado.precisa_trocar_senha:
                return redirect("/trocar_senha")

            return redirect("/")

        erro = "Usuário ou senha inválidos"

    return render_template(

        "login.html",

        erro=erro if 'erro' in locals()
        else None

    )

@app.route("/logout")
def logout():

    session.clear()

    return redirect("/login")

# =========================
# TROCAR SENHA
# =========================

@app.route(
    "/trocar_senha",
    methods=["GET", "POST"]
)
def trocar_senha():

    usuario = session.get("usuario")

    if not usuario:

        return redirect("/login")

    if request.method == "POST":

        nova_senha = request.form.get(
            "nova_senha"
        )

        confirmar_senha = request.form.get(
            "confirmar_senha"
        )

        if nova_senha != confirmar_senha:

            flash(
                "As senhas não coincidem.",
                "error"
            )

            return redirect("/trocar_senha")

        senha_hash = generate_password_hash(
            nova_senha
        )

        db.session.execute(

            db.text("""

                UPDATE usuarios

                SET

                    senha = :senha,

                    precisa_trocar_senha = FALSE

                WHERE id = :id

            """),

            {
                "senha": senha_hash,
                "id": usuario["id"]
            }

        )

        db.session.commit()

        flash(
            "Senha alterada com sucesso!",
            "success"
        )

        return redirect("/")

    return render_template(
        "trocar_senha.html",
        usuario=usuario
    )


# =========================
# DASHBOARD
# =========================

@app.route("/")
def dashboard():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    chamados = db.session.execute(text("""

        SELECT *
        FROM chamados
        ORDER BY id DESC

    """)).fetchall()

    abertos = len([
        c for c in chamados
        if c.status == "aberto"
    ])

    andamento = len([
        c for c in chamados
        if c.status == "em_andamento"
    ])

    resolvidos = len([
        c for c in chamados
        if c.status == "Finalizado"
    ])

    ranking_colaborador = Counter([
        c.usuario or "Desconhecido"
        for c in chamados
    ])

    ranking_setor = Counter([
        c.setor or "Não informado"
        for c in chamados
    ])

    problemas = Counter()

    for c in chamados:

        categoria = (
            c.categoria
            or
            "Não categorizado"
        )

        problemas[categoria] += 1

    if usuario["tipo"] == "ti":

        # =========================
        # KANBAN
        # =========================

        tarefas_pendentes = db.session.execute(

            text("""

                SELECT COUNT(*)

                FROM tarefas

                WHERE LOWER(status) = 'pendente'

            """)

        ).scalar()

        tarefas_andamento = db.session.execute(

            text("""

                SELECT COUNT(*)

                FROM tarefas

                WHERE LOWER(status) = 'andamento'

            """)

        ).scalar()

        tarefas_aguardando = db.session.execute(

            text("""

                SELECT COUNT(*)

                FROM tarefas

                WHERE LOWER(status) = 'aguardando'

            """)

        ).scalar()

        tarefas_criticas = db.session.execute(

            text("""

                SELECT *

                FROM tarefas

                WHERE LOWER(prioridade) = 'critica'

                AND LOWER(status) != 'finalizado'

                ORDER BY id DESC

                LIMIT 5

            """)

        ).fetchall()

        # =========================
        # MÁQUINAS COM PROBLEMAS
        # =========================

        maquinas_problema = db.session.execute(

            text("""

                SELECT DISTINCT maquina

                FROM chamados

                WHERE status != 'Finalizado'
                AND maquina IS NOT NULL

                LIMIT 10

            """)

        ).fetchall()

        # =========================
        # ÚLTIMAS MOVIMENTAÇÕES
        # =========================

        movimentacoes = db.session.execute(

            text("""

                SELECT *

                FROM movimentacoes_mapa

                ORDER BY data_movimentacao DESC

                LIMIT 5

            """)

        ).fetchall()

        return render_template(

            "dashboard_ti.html",

            usuario=usuario,

            chamados=chamados,

            abertos=abertos,

            andamento=andamento,

            resolvidos=resolvidos,

            ranking_colaborador=ranking_colaborador,

            ranking_setor=ranking_setor,

            setores_json=json.dumps(
                dict(ranking_setor)
            ),

            problemas_json=json.dumps(
                dict(problemas)
            ),

            tarefas_pendentes=tarefas_pendentes,
            tarefas_andamento=tarefas_andamento,
            tarefas_aguardando=tarefas_aguardando,

            tarefas_criticas=tarefas_criticas,

            maquinas_problema=maquinas_problema,

            movimentacoes=movimentacoes

        )

    return render_template(

        "dashboard_colab.html",

        usuario=usuario,

        chamados=chamados,

        abertos=abertos,

        andamento=andamento,

        resolvidos=resolvidos

    )

# =========================
# ARQUIVAR CHAMADOS ANTIGOS
# =========================

@app.route("/arquivar_chamados_antigos")
def arquivar_chamados_antigos():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    db.session.execute(

        db.text("""

            UPDATE chamados

            SET arquivado = true

            WHERE

                LOWER(status) = 'finalizado'

                AND

                finalizado_em <=
                CURRENT_TIMESTAMP
                - INTERVAL '90 days'

                AND

                arquivado = false

        """)

    )

    db.session.commit()

    flash(

        "Chamados antigos arquivados!",

        "success"

    )

    return redirect("/chamados")
# =========================
# LISTA CHAMADOS
# =========================

@app.route("/chamados")
def chamados():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    busca = request.args.get(
        "busca",
        ""
    ).lower()

    modo = request.args.get(
        "modo",
        "ativos"
    )

    arquivado = (

        modo == "arquivados"

    )

    params = {

        "arquivado": arquivado

    }

    if usuario["tipo"] == "ti":

        lista_db = db.session.execute(

            db.text("""

                SELECT *

                FROM chamados

                WHERE arquivado = :arquivado

                ORDER BY id DESC

            """),

            params

        ).fetchall()

    else:

        lista_db = db.session.execute(

            db.text("""

                SELECT *

                FROM chamados

                WHERE

                    usuario = :usuario

                    AND

                    arquivado = :arquivado

                ORDER BY id DESC

            """),

            {

                "usuario":
                usuario["nome"],

                "arquivado":
                arquivado

            }

        ).fetchall()

    lista = []

    for c in lista_db:

        tecnicos_db = db.session.execute(

            db.text("""

                SELECT tecnico

                FROM chamados_tecnicos

                WHERE chamado_id = :id

            """),

            {
                "id": c.id
            }

        ).fetchall()

        tecnicos = [

            t.tecnico

            for t in tecnicos_db

        ]

        lista.append({

            "id": c.id,
            "usuario": c.usuario,
            "descricao": c.descricao,
            "setor": c.setor,
            "maquina": c.maquina,
            "status": c.status,
            "categoria": c.categoria,
            "solucao": c.solucao,
            "origem": c.origem,
            "tecnicos": tecnicos

        })

    if busca:

        lista = [

            c for c in lista

            if

            busca in str(
                c["id"]
            ).lower()

            or

            busca in str(
                c["descricao"]
            ).lower()

            or

            busca in str(
                c["usuario"]
            ).lower()

            or

            busca in str(
                c["maquina"]
            ).lower()

        ]

    return render_template(

        "chamados.html",

        chamados=lista,

        usuario=usuario,

        busca=busca,

        modo=modo

    )
# =========================
# DETALHE
# =========================

@app.route("/chamado/<int:id>")
def detalhe_chamado(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    chamado = db.session.execute(

        db.text("""

            SELECT *

            FROM chamados

            WHERE id = :id

        """),

        {
            "id": id
        }

    ).fetchone()

    if not chamado:
        return redirect("/chamados")

    if (

        usuario["tipo"] != "ti"

        and

        chamado.usuario != usuario["nome"]

    ):

        return redirect("/chamados")

    mensagens = db.session.execute(

        db.text("""

            SELECT *

            FROM mensagens

            WHERE chamado_id = :id

            ORDER BY id ASC

        """),

        {
            "id": id
        }

    ).fetchall()

    tecnicos_db = db.session.execute(

        db.text("""

            SELECT tecnico

            FROM chamados_tecnicos

            WHERE chamado_id = :id

        """),

        {
            "id": id
        }

    ).fetchall()

    tecnicos = [

        t.tecnico

        for t in tecnicos_db

    ]

    return render_template(

        "detalhe.html",

        chamado=chamado,

        mensagens=mensagens,

        tecnicos=tecnicos,

        usuario=usuario

    )
# =========================
# NOVO CHAMADO
# =========================

@app.route(
    "/novo",
    methods=["GET", "POST"]
)
def novo_chamado():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] == "ti":
        return redirect("/")

    if request.method == "POST":

        db.session.execute(

            db.text("""

                INSERT INTO chamados (

                    usuario,
                    descricao,
                    setor,
                    maquina,
                    status,
                    categoria,
                    solucao

                )

                VALUES (

                    :usuario,
                    :descricao,
                    :setor,
                    :maquina,
                    'aberto',
                    '',
                    ''

                )

            """),

            {

                "usuario": usuario["nome"],

                "descricao": request.form.get(
                    "descricao",
                    ""
                ),

                "setor": request.form.get(
                    "setor",
                    ""
                ),

                "maquina": request.form.get(
                    "maquina",
                    ""
                )

            }

        )

        chamado_criado = db.session.execute(

            db.text("""

                SELECT id

                FROM chamados

                WHERE usuario = :usuario

                ORDER BY id DESC

                LIMIT 1

            """),

            {
                "usuario": usuario["nome"]
            }

        ).fetchone()

        if chamado_criado:

            tecnicos = db.session.execute(

                db.text("""

                    SELECT id

                    FROM usuarios

                    WHERE tipo IN ('ti', 'administracao')
                    AND ativo = true

                """)

            ).fetchall()

            for tecnico in tecnicos:

                criar_notificacao(

                    usuario_id=tecnico.id,

                    titulo="Novo chamado aberto",

                    mensagem=f"{usuario['nome']} abriu o chamado #{chamado_criado.id}.",

                    tipo="chamado",

                    link=f"/chamado/{chamado_criado.id}"

                )

        db.session.commit()

        return redirect("/chamados")

    return render_template(
        "novo.html",
        usuario=usuario
    )
# =========================
# CHAT
# =========================

@app.route(
    "/mensagem/<int:chamado_id>",
    methods=["POST"]
)
def enviar_mensagem(chamado_id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    texto = request.form.get(
        "mensagem",
        ""
    )

    arquivo_nome = ""

    arquivo = request.files.get("arquivo")

    if arquivo and arquivo.filename != "":

        pasta = os.path.join(
            UPLOAD_FOLDER,
            f"chat_{chamado_id}"
        )

        os.makedirs(
            pasta,
            exist_ok=True
        )

        nome_seguro = secure_filename(
            arquivo.filename
        )

        caminho = os.path.join(
            pasta,
            nome_seguro
        )

        arquivo.save(caminho)

        arquivo_nome = (
            f"chat_{chamado_id}/{nome_seguro}"
        )

    db.session.execute(

        db.text("""

            INSERT INTO mensagens (

                chamado_id,
                usuario,
                mensagem,
                arquivo

            )

            VALUES (

                :chamado_id,
                :usuario,
                :mensagem,
                :arquivo

            )

        """),

        {

            "chamado_id": chamado_id,

            "usuario": usuario["nome"],

            "mensagem": texto,

            "arquivo": arquivo_nome

        }

    )

    chamado = db.session.execute(

        db.text("""

            SELECT *

            FROM chamados

            WHERE id = :id

        """),

        {
            "id": chamado_id
        }

    ).fetchone()

    if chamado:

        # =========================
        # SE COLABORADOR ENVIOU
        # NOTIFICA TODOS OS TIs
        # =========================

        if usuario["tipo"] not in [
            "ti",
            "administracao"
        ]:

            tecnicos = db.session.execute(

                db.text("""

                    SELECT id

                    FROM usuarios

                    WHERE tipo IN ('ti', 'administracao')
                    AND ativo = true

                """)

            ).fetchall()

            for tecnico in tecnicos:

                criar_notificacao(

                    usuario_id=tecnico.id,

                    titulo="Nova mensagem em chamado",

                    mensagem=f"{usuario['nome']} enviou uma mensagem no chamado #{chamado_id}.",

                    tipo="mensagem",

                    link=f"/chamado/{chamado_id}"

                )

        # =========================
        # SE TI / ADMIN ENVIOU
        # NOTIFICA O DONO DO CHAMADO
        # =========================

        else:

            dono_chamado = db.session.execute(

                db.text("""

                    SELECT id

                    FROM usuarios

                    WHERE LOWER(nome) = LOWER(:nome)

                    LIMIT 1

                """),

                {
                    "nome": chamado.usuario
                }

            ).fetchone()

            if dono_chamado:

                criar_notificacao(

                    usuario_id=dono_chamado.id,

                    titulo="Nova resposta no seu chamado",

                    mensagem=f"{usuario['nome']} respondeu o chamado #{chamado_id}.",

                    tipo="mensagem",

                    link=f"/chamado/{chamado_id}"

                )

    db.session.commit()

    return redirect(
        f"/chamado/{chamado_id}"
    )
# =========================
# CATEGORIA
# =========================

@app.route(
    "/categoria/<int:chamado_id>",
    methods=["POST"]
)
def categoria_chamado(chamado_id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    categoria = request.form.get(
        "categoria",
        ""
    )

    db.session.execute(

        db.text("""

            UPDATE chamados

            SET categoria = :categoria

            WHERE id = :id

        """),

        {

            "categoria": categoria,

            "id": chamado_id

        }

    )

    db.session.commit()

    return redirect(
        f"/chamado/{chamado_id}"
    )
# =========================
# SOLUÇÃO
# =========================

@app.route(
    "/solucao/<int:chamado_id>",
    methods=["POST"]
)
def adicionar_solucao(chamado_id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    solucao = request.form.get(
        "solucao",
        ""
    )

    db.session.execute(

        db.text("""

            UPDATE chamados

            SET solucao = :solucao

            WHERE id = :id

        """),

        {

            "solucao": solucao,

            "id": chamado_id

        }

    )

    db.session.commit()

    return redirect(
        f"/chamado/{chamado_id}"
    )
# =========================
# ASSUMIR CHAMADO
# =========================

@app.route("/assumir/<int:chamado_id>")
def assumir_chamado(chamado_id):

    usuario = session.get("usuario")

    if not usuario:

        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:

        return redirect("/")

    existe = db.session.execute(

        db.text("""

            SELECT *

            FROM chamados_tecnicos

            WHERE chamado_id = :id
            AND tecnico = :tecnico

        """),

        {
            "id": chamado_id,
            "tecnico": usuario["nome"]
        }

    ).fetchone()

    if not existe:

        db.session.execute(

            db.text("""

                INSERT INTO chamados_tecnicos (

                    chamado_id,
                    tecnico

                )

                VALUES (

                    :id,
                    :tecnico

                )

            """),

            {
                "id": chamado_id,
                "tecnico": usuario["nome"]
            }

        )

    db.session.execute(

        db.text("""

            UPDATE chamados

            SET status = 'em_andamento'

            WHERE id = :id

        """),

        {
            "id": chamado_id
        }

    )

    chamado = db.session.execute(

        db.text("""

            SELECT *

            FROM chamados

            WHERE id = :id

        """),

        {
            "id": chamado_id
        }

    ).fetchone()

    usuario_chamado = db.session.execute(

        db.text("""

            SELECT id, email

            FROM usuarios

            WHERE nome = :nome

        """),

        {
            "nome": chamado.usuario
        }

    ).fetchone()

    email_usuario = None

    if usuario_chamado:

        email_usuario = usuario_chamado.email

        # =========================
        # NOTIFICAÇÃO
        # =========================

        criar_notificacao(

            usuario_id=usuario_chamado.id,

            titulo="Chamado assumido",

            mensagem=f"{usuario['nome']} assumiu seu chamado #{chamado.id}.",

            tipo="chamado",

            link=f"/chamado/{chamado.id}"

        )

    if email_usuario:

        try:

            msg = Message(
                subject="Seu chamado foi assumido",
                recipients=[email_usuario]
            )

            msg.sender = (
                "Sistema TI",
                "demo@example.com"
            )

            msg.html = f"""

            <div style="
                background:#f4f7fb;
                padding:40px;
                font-family:Arial,sans-serif;
            ">

                <div style="
                    max-width:600px;
                    margin:auto;
                    background:white;
                    border-radius:16px;
                    overflow:hidden;
                    border:1px solid #dbe4ee;
                    box-shadow:0 4px 20px rgba(0,0,0,0.08);
                ">

                    <div style="
                        background:#2563eb;
                        padding:25px;
                        color:white;
                    ">

                        <h1 style="
                            margin:0;
                            font-size:24px;
                        ">
                            Sistema de Chamados TI
                        </h1>

                    </div>

                    <div style="padding:30px;">

                        <h2 style="
                            color:#0f172a;
                            margin-top:0;
                        ">
                            Chamado em atendimento
                        </h2>

                        <p style="
                            color:#334155;
                            line-height:1.7;
                            font-size:15px;
                        ">
                            Olá,
                        </p>

                        <p style="
                            color:#334155;
                            line-height:1.7;
                            font-size:15px;
                        ">
                            Seu chamado foi assumido
                            e já está em atendimento.
                        </p>

                        <div style="
                            background:#f8fafc;
                            border:1px solid #e2e8f0;
                            border-radius:10px;
                            padding:20px;
                            margin:25px 0;
                        ">

                            <p style="margin:0 0 10px 0;">
                                <strong>Status:</strong>
                                Em andamento
                            </p>

                            <p style="margin:0 0 10px 0;">
                                <strong>Técnico:</strong>
                                {session["usuario"]["nome"]}
                            </p>

                            <p style="margin:0;">
                                <strong>Chamado:</strong>
                                #{chamado.id}
                            </p>

                        </div>

                        <p style="
                            color:#64748b;
                            font-size:14px;
                            line-height:1.6;
                        ">
                            Em breve a equipe de TI
                            irá analisar o problema.
                        </p>

                    </div>

                    <div style="
                        background:#f8fafc;
                        padding:20px;
                        border-top:1px solid #e2e8f0;
                        text-align:center;
                    ">

                        <p style="
                            margin:0;
                            color:#94a3b8;
                            font-size:12px;
                        ">
                            Mensagem automática • Não responda este email
                        </p>

                    </div>

                </div>

            </div>

            """

            mail.send(msg)

        except Exception as e:

            print(e)

    db.session.commit()

    return redirect(f"/chamado/{chamado_id}")
# =========================
# FINALIZAR CHAMADO
# =========================

@app.route("/finalizar/<int:id>")
def finalizar(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    chamado = db.session.execute(

        db.text("""

            SELECT *

            FROM chamados

            WHERE id = :id

        """),

        {
            "id": id
        }

    ).fetchone()

    if not chamado:

        flash(
            "Chamado não encontrado.",
            "error"
        )

        return redirect("/chamados")

    db.session.execute(

        db.text("""

            UPDATE chamados

            SET

                status = 'Finalizado',

                finalizado_em = CURRENT_TIMESTAMP

            WHERE id = :id

        """),

        {
            "id": id
        }

    )

    usuario_chamado = db.session.execute(

        db.text("""

            SELECT id

            FROM usuarios

            WHERE LOWER(nome)=LOWER(:nome)

            LIMIT 1

        """),

        {
            "nome": chamado.usuario
        }

    ).fetchone()

    if usuario_chamado:

        criar_notificacao(

            usuario_id=usuario_chamado.id,

            titulo="Chamado finalizado",

            mensagem=f"{usuario['nome']} finalizou seu chamado #{id}.",

            tipo="chamado",

            link=f"/chamado/{id}"

        )

    db.session.commit()

    flash(
        "Chamado finalizado com sucesso!",
        "success"
    )

    return redirect("/chamados")

# =========================
# EXCLUIR
# =========================

@app.route(
    "/excluir/<int:chamado_id>",
    methods=["GET", "POST"]
)
def excluir_chamado(chamado_id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    chamado = db.session.execute(text("""

        SELECT *
        FROM chamados
        WHERE id = :id

    """), {
        "id": chamado_id
    }).fetchone()

    if not chamado:
        return redirect("/chamados")

    if request.method == "POST":

        tecnicos_vinculados = db.session.execute(text("""

            SELECT tecnico
            FROM chamados_tecnicos
            WHERE chamado_id = :id

        """), {
            "id": chamado_id
        }).fetchall()

        usuario_dono = db.session.execute(text("""

            SELECT id
            FROM usuarios
            WHERE LOWER(nome) = LOWER(:nome)
            LIMIT 1

        """), {
            "nome": chamado.usuario
        }).fetchone()

        if usuario_dono:

            criar_notificacao(

                usuario_id=usuario_dono.id,

                titulo="Chamado excluído",

                mensagem=f"O chamado #{chamado_id} foi excluído por {usuario['nome']}.",

                tipo="chamado",

                link="/chamados"

            )

        for tecnico in tecnicos_vinculados:

            tecnico_db = db.session.execute(text("""

                SELECT id
                FROM usuarios
                WHERE LOWER(nome) = LOWER(:nome)
                LIMIT 1

            """), {
                "nome": tecnico.tecnico
            }).fetchone()

            if tecnico_db:

                criar_notificacao(

                    usuario_id=tecnico_db.id,

                    titulo="Chamado excluído",

                    mensagem=f"O chamado #{chamado_id}, vinculado a você, foi excluído por {usuario['nome']}.",

                    tipo="chamado",

                    link="/chamados"

                )

        # REMOVE MENSAGENS RELACIONADAS

        db.session.execute(text("""

            DELETE FROM mensagens

            WHERE chamado_id = :id

        """), {
            "id": chamado_id
        })

        # REMOVE TÉCNICOS RELACIONADOS

        db.session.execute(text("""

            DELETE FROM chamados_tecnicos

            WHERE chamado_id = :id

        """), {
            "id": chamado_id
        })

        # REMOVE CHAMADO

        db.session.execute(text("""

            DELETE FROM chamados

            WHERE id = :id

        """), {
            "id": chamado_id
        })

        db.session.commit()

        flash(
            "Chamado excluído com sucesso!",
            "success"
        )

        return redirect("/chamados")

    return render_template(

        "confirmar_exclusao.html",

        usuario=usuario,

        chamado=chamado

    )
# =========================
# HISTÓRICO
# =========================

@app.route("/historico")
def historico():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    busca = request.args.get(
        "busca",
        ""
    ).lower()

    resultados = []

    tecnicos = set()

    maquinas = set()

    if busca:

        resultados = db.session.execute(

            text("""

                SELECT *

                FROM chamados

                WHERE

                    LOWER(maquina) LIKE :busca

                    OR

                    LOWER(usuario) LIKE :busca

                ORDER BY id DESC

            """),

            {
                "busca": f"%{busca}%"
            }

        ).fetchall()

        for c in resultados:

            maquinas.add(c.maquina)

            tecnicos_db = db.session.execute(

                text("""

                    SELECT tecnico

                    FROM chamados_tecnicos

                    WHERE chamado_id = :id

                """),

                {
                    "id": c.id
                }

            ).fetchall()

            for t in tecnicos_db:

                tecnicos.add(t.tecnico)

    return render_template(

        "historico.html",

        usuario=usuario,

        resultados=resultados,

        busca=busca,

        tecnicos=tecnicos,

        maquinas=maquinas

    )

# =========================
# ATENDIMENTOS TI
# =========================

@app.route("/atendimentos")
def atendimentos():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    busca = request.args.get(
        "busca",
        ""
    ).strip()

    mes = request.args.get(
        "mes",
        ""
    ).strip()

    ano = request.args.get(
        "ano",
        ""
    ).strip()

    todos = request.args.get(
        "todos",
        ""
    ) == "1"

    tecnico = request.args.get(
        "tecnico",
        ""
    ).strip()

    setor = request.args.get(
        "setor",
        ""
    ).strip()

    categoria = request.args.get(
        "categoria",
        ""
    ).strip()

    # Abre por padrão no mês atual,
    # evitando o mega listão
    if not todos and not mes and not ano:

        mes = str(datetime.now().month)
        ano = str(datetime.now().year)

    query = """

        SELECT *

        FROM atendimentos_ti

        WHERE 1=1

    """

    params = {}

    if busca:

        query += """

            AND (

                LOWER(nome_colaborador)
                LIKE LOWER(:busca)

                OR

                LOWER(problema)
                LIKE LOWER(:busca)

                OR

                LOWER(solucao)
                LIKE LOWER(:busca)

            )

        """

        params["busca"] = f"%{busca}%"

    if mes and not todos:

        query += """

            AND EXTRACT(MONTH FROM data_atendimento) = :mes

        """

        params["mes"] = int(mes)

    if ano and not todos:

        query += """

            AND EXTRACT(YEAR FROM data_atendimento) = :ano

        """

        params["ano"] = int(ano)

    if tecnico:

        query += """

            AND LOWER(tecnico) = LOWER(:tecnico)

        """

        params["tecnico"] = tecnico

    if setor:

        query += """

            AND LOWER(setor) LIKE LOWER(:setor)

        """

        params["setor"] = f"%{setor}%"

    if categoria:

        query += """

            AND LOWER(categoria) = LOWER(:categoria)

        """

        params["categoria"] = categoria

    query += """

        ORDER BY data_atendimento DESC, id DESC

    """

    lista = db.session.execute(
        db.text(query),
        params
    ).fetchall()
    # =========================
    # MINI DASHBOARD ATENDIEMNTOS
    # =========================
    contador_categoria = Counter([
        a.categoria or "Não informado"
        for a in lista
    ])
    contador_setor = Counter([
        a.setor or "Não informado"
        for a in lista
    ])
    total_periodo = len(lista)
    
    contador_dia = Counter([
        a.data_atendimento.strftime("%d-%m")
        for a in lista
        if a.data_atendimento
    ])
    total_perdido = len(lista)
    categoria_top = (
        contador_categoria.most_common(1)[0]
        if contador_categoria
        else ("-", 0)
    )
    setor_top = (
        contador_setor.most_common(1)[0]
        if contador_setor
        else ("-", 0)
    )
    max_categoria = max(
        contador_categoria.values(),
        default=1
    )
    max_setor = max(
        contador_setor.values(),
        default=1
    )


    tecnicos = db.session.execute(

        db.text("""

            SELECT nome

            FROM usuarios

            WHERE tipo = 'ti'
            AND ativo = true

            ORDER BY nome

        """)

    ).fetchall()

    total_atendimentos = db.session.execute(

        db.text("""

            SELECT COUNT(*)

            FROM atendimentos_ti

        """)

    ).scalar()

    anos_db = db.session.execute(

        db.text("""

            SELECT DISTINCT
                EXTRACT(YEAR FROM data_atendimento)::INT AS ano

            FROM atendimentos_ti

            ORDER BY ano DESC

        """)

    ).fetchall()

    anos_disponiveis = [

        a.ano

        for a in anos_db

    ]

    if not anos_disponiveis:

        anos_disponiveis = [
            datetime.now().year
        ]

    meses = [
        ("1", "Jan"),
        ("2", "Fev"),
        ("3", "Mar"),
        ("4", "Abr"),
        ("5", "Mai"),
        ("6", "Jun"),
        ("7", "Jul"),
        ("8", "Ago"),
        ("9", "Set"),
        ("10", "Out"),
        ("11", "Nov"),
        ("12", "Dez")
    ]

    ano_base = ano or str(datetime.now().year)

    return render_template(
        "atendimentos.html",
        usuario=usuario,
        atendimentos=lista,
        tecnicos=tecnicos,
        busca=busca,
        mes=mes,
        ano=ano,
        ano_base=ano_base,
        todos=todos,
        meses=meses,
        anos_disponiveis=anos_disponiveis,
        tecnico=tecnico,
        setor=setor,
        categoria=categoria,
        total_atendimentos=total_atendimentos,
        total_periodo=total_periodo,
        categoria_top=categoria_top,
        setor_top=setor_top,
        contador_categoria=contador_categoria,
        contador_setor=contador_setor,
        contador_dia=contador_dia,
        max_categoria=max_categoria,
        max_setor=max_setor
    )


    
# =========================
# NOVO ATENDIMENTO
# =========================

@app.route(
    "/novo_atendimento",
    methods=["POST"]
)
def novo_atendimento():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    db.session.execute(

        db.text("""

            INSERT INTO atendimentos_ti (

                nome_colaborador,
                setor,
                data_atendimento,
                problema,
                solucao,
                categoria,
                tecnico,
                criado_por

            )

            VALUES (

                :nome_colaborador,
                :setor,
                :data_atendimento,
                :problema,
                :solucao,
                :categoria,
                :tecnico,
                :criado_por

            )

        """),

        {
            "nome_colaborador":
            request.form.get("nome_colaborador", ""),

            "setor":
            request.form.get("setor", ""),

            "data_atendimento":
            request.form.get("data_atendimento"),

            "problema":
            request.form.get("problema", ""),

            "solucao":
            request.form.get("solucao", ""),

            "categoria":
            request.form.get("categoria", ""),

            "tecnico":
            request.form.get("tecnico", ""),

            "criado_por":
            usuario["nome"]
        }

    )

    db.session.commit()

    flash(
        "Atendimento registrado com sucesso!",
        "success"
    )

    return redirect("/atendimentos")


# =========================
# EDITAR ATENDIMENTO
# =========================

@app.route(
    "/editar_atendimento/<int:id>",
    methods=["GET", "POST"]
)
def editar_atendimento(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    atendimento = db.session.execute(

        db.text("""

            SELECT *

            FROM atendimentos_ti

            WHERE id = :id

        """),

        {
            "id": id
        }

    ).fetchone()

    if not atendimento:

        flash(
            "Atendimento não encontrado.",
            "error"
        )

        return redirect("/atendimentos")

    tecnicos = db.session.execute(

        db.text("""

            SELECT nome

            FROM usuarios

            WHERE tipo = 'ti'
            AND ativo = true

            ORDER BY nome

        """)

    ).fetchall()

    if request.method == "POST":

        db.session.execute(

            db.text("""

                UPDATE atendimentos_ti

                SET

                    nome_colaborador = :nome_colaborador,
                    setor = :setor,
                    data_atendimento = :data_atendimento,
                    problema = :problema,
                    solucao = :solucao,
                    categoria = :categoria,
                    tecnico = :tecnico

                WHERE id = :id

            """),

            {
                "id":
                id,

                "nome_colaborador":
                request.form.get("nome_colaborador", ""),

                "setor":
                request.form.get("setor", ""),

                "data_atendimento":
                request.form.get("data_atendimento"),

                "problema":
                request.form.get("problema", ""),

                "solucao":
                request.form.get("solucao", ""),

                "categoria":
                request.form.get("categoria", ""),

                "tecnico":
                request.form.get("tecnico", "")
            }

        )

        db.session.commit()

        flash(
            "Atendimento atualizado com sucesso!",
            "success"
        )

        return redirect("/atendimentos")

    return render_template(

        "editar_atendimento.html",

        usuario=usuario,

        atendimento=atendimento,

        tecnicos=tecnicos

    )


# =========================
# EXCLUIR ATENDIMENTO
# =========================

@app.route("/excluir_atendimento/<int:id>")
def excluir_atendimento(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    db.session.execute(

        db.text("""

            DELETE FROM atendimentos_ti

            WHERE id = :id

        """),

        {
            "id": id
        }

    )

    db.session.commit()

    flash(
        "Atendimento excluído.",
        "success"
    )

    return redirect("/atendimentos")


# =========================
# EXPORTAR ATENDIMENTOS XLSX
# =========================

@app.route("/exportar_atendimentos_xlsx")
def exportar_atendimentos_xlsx():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    busca = request.args.get(
        "busca",
        ""
    ).strip()

    mes = request.args.get(
        "mes",
        ""
    ).strip()

    ano = request.args.get(
        "ano",
        ""
    ).strip()

    todos = request.args.get(
        "todos",
        ""
    ) == "1"

    tecnico = request.args.get(
        "tecnico",
        ""
    ).strip()

    setor = request.args.get(
        "setor",
        ""
    ).strip()

    categoria = request.args.get(
        "categoria",
        ""
    ).strip()

    if not todos and not mes and not ano:

        mes = str(datetime.now().month)
        ano = str(datetime.now().year)

    query = """

        SELECT *

        FROM atendimentos_ti

        WHERE 1=1

    """

    params = {}

    if busca:

        query += """

            AND (

                LOWER(nome_colaborador)
                LIKE LOWER(:busca)

                OR

                LOWER(problema)
                LIKE LOWER(:busca)

                OR

                LOWER(solucao)
                LIKE LOWER(:busca)

            )

        """

        params["busca"] = f"%{busca}%"

    if mes and not todos:

        query += """

            AND EXTRACT(MONTH FROM data_atendimento) = :mes

        """

        params["mes"] = int(mes)

    if ano and not todos:

        query += """

            AND EXTRACT(YEAR FROM data_atendimento) = :ano

        """

        params["ano"] = int(ano)

    if tecnico:

        query += """

            AND LOWER(tecnico) = LOWER(:tecnico)

        """

        params["tecnico"] = tecnico

    if setor:

        query += """

            AND LOWER(setor) LIKE LOWER(:setor)

        """

        params["setor"] = f"%{setor}%"

    if categoria:

        query += """

            AND LOWER(categoria) = LOWER(:categoria)

        """

        params["categoria"] = categoria

    query += """

        ORDER BY data_atendimento DESC, id DESC

    """

    atendimentos = db.session.execute(
        db.text(query),
        params
    ).fetchall()

    meses_nome = {
        "1": "Janeiro",
        "2": "Fevereiro",
        "3": "Março",
        "4": "Abril",
        "5": "Maio",
        "6": "Junho",
        "7": "Julho",
        "8": "Agosto",
        "9": "Setembro",
        "10": "Outubro",
        "11": "Novembro",
        "12": "Dezembro"
    }

    if todos:

        periodo_texto = "Todos os períodos"
        periodo_arquivo = "todos"

    elif mes and ano:

        periodo_texto = f"{meses_nome.get(mes, mes)} / {ano}"
        periodo_arquivo = f"{ano}_{str(mes).zfill(2)}"

    else:

        periodo_texto = "Período filtrado"
        periodo_arquivo = "filtrado"

    wb = Workbook()

    ws = wb.active

    ws.title = "Atendimentos TI"

    ws.sheet_view.showGridLines = False

    # =========================
    # CORES E ESTILOS
    # =========================

    azul_noite = "081426"
    azul_profundo = "0F2747"
    azul_vyron = "2563EB"
    azul_claro = "DBEAFE"
    branco = "FFFFFF"
    cinza_borda = "CBD5E1"
    cinza_texto = "64748B"
    cinza_linha = "F8FAFC"
    cinza_linha_2 = "EEF2F7"
    verde_claro = "DCFCE7"
    vermelho_claro = "FEE2E2"
    amarelo_claro = "FEF3C7"
    roxo_claro = "EDE9FE"

    fill_titulo = PatternFill(
        "solid",
        fgColor=azul_noite
    )

    fill_subtitulo = PatternFill(
        "solid",
        fgColor=azul_profundo
    )

    fill_header = PatternFill(
        "solid",
        fgColor=azul_vyron
    )

    fill_kpi = PatternFill(
        "solid",
        fgColor=azul_claro
    )

    fill_linha = PatternFill(
        "solid",
        fgColor=cinza_linha
    )

    fill_linha_2 = PatternFill(
        "solid",
        fgColor=cinza_linha_2
    )

    fonte_titulo = Font(
        color=branco,
        bold=True,
        size=20
    )

    fonte_subtitulo = Font(
        color=branco,
        bold=True,
        size=11
    )

    fonte_header = Font(
        color=branco,
        bold=True,
        size=11
    )

    fonte_kpi_label = Font(
        color=cinza_texto,
        bold=True,
        size=10
    )

    fonte_kpi_valor = Font(
        color=azul_noite,
        bold=True,
        size=18
    )

    fonte_padrao = Font(
        color="0F172A",
        size=10
    )

    borda = Border(
        left=Side(style="thin", color=cinza_borda),
        right=Side(style="thin", color=cinza_borda),
        top=Side(style="thin", color=cinza_borda),
        bottom=Side(style="thin", color=cinza_borda)
    )

    centro = Alignment(
        horizontal="center",
        vertical="center"
    )

    esquerda = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    # =========================
    # CABEÇALHO
    # =========================

    ws.merge_cells("A1:H2")

    ws["A1"] = "Relatório de Atendimentos TI"

    for row in range(1, 3):

        for col in range(1, 9):

            celula = ws.cell(
                row=row,
                column=col
            )

            celula.fill = fill_titulo
            celula.border = borda

    ws["A1"].font = fonte_titulo
    ws["A1"].alignment = centro

    ws.merge_cells("A3:H3")

    ws["A3"] = (
        f"Vyron ITSM • {periodo_texto} • "
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )

    for col in range(1, 9):

        celula = ws.cell(
            row=3,
            column=col
        )

        celula.fill = fill_subtitulo
        celula.border = borda

    ws["A3"].font = fonte_subtitulo
    ws["A3"].alignment = centro

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 24

        # =========================
    # KPIs
    # =========================

    total_atendimentos = len(atendimentos)

    setores = set()
    categorias = set()
    tecnicos = set()

    for a in atendimentos:

        if a.setor:
            setores.add(a.setor)

        if a.categoria:
            categorias.add(a.categoria)

        if a.tecnico:
            tecnicos.add(a.tecnico)

    resumo = [
        ("Atendimentos", total_atendimentos),
        ("Setores", len(setores)),
        ("Categorias", len(categorias)),
        ("Técnicos", len(tecnicos))
    ]

    blocos = [
        {
            "label_range": "A5:B5",
            "valor_range": "A6:B6",
            "label_cell": "A5",
            "valor_cell": "A6",
            "col_inicio": 1,
            "col_fim": 2
        },
        {
            "label_range": "C5:D5",
            "valor_range": "C6:D6",
            "label_cell": "C5",
            "valor_cell": "C6",
            "col_inicio": 3,
            "col_fim": 4
        },
        {
            "label_range": "E5:F5",
            "valor_range": "E6:F6",
            "label_cell": "E5",
            "valor_cell": "E6",
            "col_inicio": 5,
            "col_fim": 6
        },
        {
            "label_range": "G5:H5",
            "valor_range": "G6:H6",
            "label_cell": "G5",
            "valor_cell": "G6",
            "col_inicio": 7,
            "col_fim": 8
        }
    ]

    for i, (label, valor) in enumerate(resumo):

        bloco = blocos[i]

        ws.merge_cells(bloco["label_range"])
        ws.merge_cells(bloco["valor_range"])

        for row in range(5, 7):

            for col in range(
                bloco["col_inicio"],
                bloco["col_fim"] + 1
            ):

                celula = ws.cell(
                    row=row,
                    column=col
                )

                celula.fill = fill_kpi
                celula.border = borda
                celula.alignment = centro

        ws[bloco["label_cell"]] = label
        ws[bloco["label_cell"]].font = fonte_kpi_label
        ws[bloco["label_cell"]].alignment = centro

        ws[bloco["valor_cell"]] = valor
        ws[bloco["valor_cell"]].font = fonte_kpi_valor
        ws[bloco["valor_cell"]].alignment = centro

    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 30
    # =========================
    # TABELA
    # =========================

    linha_header = 9

    headers = [
        "ID",
        "Data",
        "Nome",
        "Setor",
        "Problema",
        "Solução",
        "Categoria",
        "Técnico"
    ]

    for col, header in enumerate(headers, start=1):

        celula = ws.cell(
            row=linha_header,
            column=col,
            value=header
        )

        celula.fill = fill_header
        celula.font = fonte_header
        celula.border = borda
        celula.alignment = centro

    ws.row_dimensions[linha_header].height = 28

    linha = linha_header + 1

    categoria_fills = {
        "Hardware": PatternFill("solid", fgColor=vermelho_claro),
        "Software": PatternFill("solid", fgColor=azul_claro),
        "Acesso": PatternFill("solid", fgColor=amarelo_claro),
        "Rede": PatternFill("solid", fgColor=verde_claro),
        "Telefonia": PatternFill("solid", fgColor=roxo_claro),
        "Outro": PatternFill("solid", fgColor=cinza_linha_2)
    }

    for a in atendimentos:

        data_formatada = ""

        if a.data_atendimento:

            data_formatada = a.data_atendimento.strftime(
                "%d/%m/%Y"
            )

        dados = [
            a.id,
            data_formatada,
            a.nome_colaborador,
            a.setor,
            a.problema,
            a.solucao,
            a.categoria,
            a.tecnico
        ]

        fill_atual = (
            fill_linha
            if linha % 2 == 0
            else fill_linha_2
        )

        for col, valor in enumerate(dados, start=1):

            celula = ws.cell(
                row=linha,
                column=col,
                value=valor or "-"
            )

            celula.fill = fill_atual
            celula.border = borda
            celula.font = fonte_padrao
            celula.alignment = esquerda

            if col in [
                1,
                2,
                4,
                7,
                8
            ]:

                celula.alignment = centro

            if col == 7:

                categoria_nome = str(valor or "").strip()

                if categoria_nome in categoria_fills:

                    celula.fill = categoria_fills[categoria_nome]

                celula.font = Font(
                    color="0F172A",
                    bold=True,
                    size=10
                )

        ws.row_dimensions[linha].height = 36

        linha += 1

    ultima_linha = linha - 1

    if ultima_linha >= linha_header + 1:

        tabela = Table(
            displayName="TabelaAtendimentosTI",
            ref=f"A{linha_header}:H{ultima_linha}"
        )

        estilo_tabela = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False
        )

        tabela.tableStyleInfo = estilo_tabela

        ws.add_table(tabela)

    ws.auto_filter.ref = f"A{linha_header}:H{max(ultima_linha, linha_header)}"

    ws.freeze_panes = "A10"

    larguras = {
        "A": 8,
        "B": 14,
        "C": 26,
        "D": 16,
        "E": 44,
        "F": 50,
        "G": 16,
        "H": 22
    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[coluna].width = largura

    # =========================
    # ABA RESUMOS
    # =========================

    ws2 = wb.create_sheet("Resumos")

    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:E2")

    ws2["A1"] = "Resumo dos Atendimentos TI"

    for row in range(1, 3):

        for col in range(1, 6):

            celula = ws2.cell(
                row=row,
                column=col
            )

            celula.fill = fill_titulo
            celula.border = borda

    ws2["A1"].font = fonte_titulo
    ws2["A1"].alignment = centro

    def escrever_resumo(titulo, contador, linha_inicio):

        ws2.merge_cells(
            start_row=linha_inicio,
            start_column=1,
            end_row=linha_inicio,
            end_column=3
        )

        ws2.cell(
            row=linha_inicio,
            column=1,
            value=titulo
        )

        ws2.cell(
            row=linha_inicio,
            column=1
        ).fill = fill_subtitulo

        ws2.cell(
            row=linha_inicio,
            column=1
        ).font = fonte_subtitulo

        ws2.cell(
            row=linha_inicio,
            column=1
        ).alignment = centro

        for col in range(1, 4):

            ws2.cell(
                row=linha_inicio,
                column=col
            ).fill = fill_subtitulo

            ws2.cell(
                row=linha_inicio,
                column=col
            ).border = borda

        headers_resumo = [
            "Item",
            "Quantidade",
            "%"
        ]

        for col, header in enumerate(headers_resumo, start=1):

            celula = ws2.cell(
                row=linha_inicio + 1,
                column=col,
                value=header
            )

            celula.fill = fill_header
            celula.font = fonte_header
            celula.border = borda
            celula.alignment = centro

        linha_resumo = linha_inicio + 2

        total = sum(
            contador.values()
        )

        for chave, quantidade in contador.most_common():

            percentual = (
                quantidade / total
                if total
                else 0
            )

            ws2.cell(
                row=linha_resumo,
                column=1,
                value=chave or "Não informado"
            )

            ws2.cell(
                row=linha_resumo,
                column=2,
                value=quantidade
            )

            ws2.cell(
                row=linha_resumo,
                column=3,
                value=percentual
            )

            ws2.cell(
                row=linha_resumo,
                column=3
            ).number_format = "0.00%"

            for col in range(1, 4):

                celula = ws2.cell(
                    row=linha_resumo,
                    column=col
                )

                celula.border = borda
                celula.alignment = centro

            linha_resumo += 1

        return linha_resumo + 2

    contador_categoria = Counter([
        a.categoria or "Não informado"
        for a in atendimentos
    ])

    contador_setor = Counter([
        a.setor or "Não informado"
        for a in atendimentos
    ])

    contador_tecnico = Counter([
        a.tecnico or "Não informado"
        for a in atendimentos
    ])

    proxima_linha = escrever_resumo(
        "Atendimentos por Categoria",
        contador_categoria,
        5
    )

    proxima_linha = escrever_resumo(
        "Atendimentos por Setor",
        contador_setor,
        proxima_linha
    )

    escrever_resumo(
        "Atendimentos por Técnico",
        contador_tecnico,
        proxima_linha
    )

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 14

    # =========================
    # CONFIGURAÇÃO DE IMPRESSÃO
    # =========================

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{linha_header}:{linha_header}"

    # =========================
    # EXPORTAÇÃO
    # =========================

    arquivo = BytesIO()

    wb.save(arquivo)

    arquivo.seek(0)

    return send_file(

        arquivo,

        as_attachment=True,

        download_name=f"atendimentos_ti_{periodo_arquivo}.xlsx",

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )
# =========================
# IMPORTAR ATENDIMENTOS XLSX
# =========================

@app.route(
    "/importar_atendimentos",
    methods=["POST"]
)
def importar_atendimentos():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    arquivo = request.files.get("arquivo")

    tecnico_padrao = request.form.get(
        "tecnico_padrao",
        usuario["nome"]
    )

    if not arquivo or arquivo.filename == "":

        flash(
            "Nenhum arquivo selecionado.",
            "error"
        )

        return redirect("/atendimentos")

    def normalizar_coluna(nome):

        nome = unicodedata.normalize(
            "NFKD",
            str(nome)
        )

        nome = nome.encode(
            "ASCII",
            "ignore"
        ).decode(
            "ASCII"
        )

        return nome.strip().lower()

    def detectar_mes_da_aba(nome_aba):

        nome = normalizar_coluna(nome_aba)

        meses = {
            "janeiro": 1,
            "fevereiro": 2,
            "marco": 3,
            "abril": 4,
            "maio": 5,
            "junho": 6,
            "julho": 7,
            "agosto": 8,
            "setembro": 9,
            "outubro": 10,
            "novembro": 11,
            "dezembro": 12
        }

        for mes_nome, mes_numero in meses.items():

            if mes_nome in nome:
                return mes_numero

        return None

    def converter_data_atendimento(valor, mes_referencia=None):

        if pd.isna(valor):
            return None

        data_convertida = None

        # Quando o Excel já entrega como data
        if isinstance(valor, datetime):

            data_convertida = valor.date()

        elif isinstance(valor, date):

            data_convertida = valor

        else:

            texto = str(valor).strip()

            formatos = [
                "%d/%m/%Y",
                "%d-%m-%Y",
                "%Y-%m-%d",
                "%d/%m/%y",
                "%d-%m-%y"
            ]

            for formato in formatos:

                try:

                    data_convertida = datetime.strptime(
                        texto,
                        formato
                    ).date()

                    break

                except ValueError:

                    pass

            if not data_convertida:

                try:

                    data_convertida = pd.to_datetime(
                        texto,
                        dayfirst=True
                    ).date()

                except Exception:

                    return None

        # =========================
        # CORREÇÃO DE DATA INVERTIDA
        # =========================

        if mes_referencia:

            if (
                data_convertida.month != mes_referencia
                and
                data_convertida.day == mes_referencia
            ):

                try:

                    data_convertida = date(
                        data_convertida.year,
                        mes_referencia,
                        data_convertida.month
                    )

                except ValueError:

                    pass

        return data_convertida

    try:

        planilhas = pd.read_excel(
            arquivo,
            sheet_name=None
        )

        total_importados = 0
        total_ignorados = 0
        total_abas = len(planilhas)

        for nome_aba, df in planilhas.items():

            mes_referencia = detectar_mes_da_aba(
                nome_aba
            )

            df.columns = [

                normalizar_coluna(col)

                for col in df.columns

            ]

            print(
                f"Aba importada: {nome_aba}"
            )

            print(
                f"Mês detectado na aba: {mes_referencia}"
            )

            print(
                f"Colunas encontradas: {list(df.columns)}"
            )

            for _, linha in df.iterrows():

                nome_colaborador = linha.get("nome")
                setor = linha.get("setor")
                data_valor = linha.get("data")
                problema = linha.get("problema")
                solucao = linha.get("solucao")
                categoria = linha.get("categoria")

                if (
                    pd.isna(nome_colaborador)
                    or
                    pd.isna(problema)
                    or
                    pd.isna(data_valor)
                ):

                    total_ignorados += 1

                    continue

                data_atendimento = converter_data_atendimento(
                    data_valor,
                    mes_referencia
                )

                if not data_atendimento:

                    total_ignorados += 1

                    continue

                tecnico_linha = linha.get(
                    "tecnico",
                    tecnico_padrao
                )

                if pd.isna(tecnico_linha):
                    tecnico_linha = tecnico_padrao

                db.session.execute(

                    db.text("""

                        INSERT INTO atendimentos_ti (

                            nome_colaborador,
                            setor,
                            data_atendimento,
                            problema,
                            solucao,
                            categoria,
                            tecnico,
                            criado_por

                        )

                        VALUES (

                            :nome_colaborador,
                            :setor,
                            :data_atendimento,
                            :problema,
                            :solucao,
                            :categoria,
                            :tecnico,
                            :criado_por

                        )

                    """),

                    {
                        "nome_colaborador":
                        str(nome_colaborador).strip(),

                        "setor":
                        "" if pd.isna(setor) else str(setor).strip(),

                        "data_atendimento":
                        data_atendimento,

                        "problema":
                        str(problema).strip(),

                        "solucao":
                        "" if pd.isna(solucao) else str(solucao).strip(),

                        "categoria":
                        "" if pd.isna(categoria) else str(categoria).strip(),

                        "tecnico":
                        str(tecnico_linha).strip(),

                        "criado_por":
                        usuario["nome"]
                    }

                )

                total_importados += 1

        db.session.commit()

        flash(
            f"Importação concluída: {total_importados} atendimentos importados, {total_ignorados} linhas ignoradas, {total_abas} abas lidas.",
            "success"
        )

    except Exception as e:

        db.session.rollback()

        flash(
            f"Erro ao importar atendimentos: {e}",
            "error"
        )

        print(
            "ERRO IMPORTAÇÃO ATENDIMENTOS:"
        )

        print(e)

    return redirect("/atendimentos")
# =========================
# EXPORTAR PDF
# =========================

@app.route("/exportar/pdf")
def exportar_pdf():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    chamados = db.session.execute(

        text("""

            SELECT *

            FROM chamados

            ORDER BY id DESC

        """)

    ).fetchall()

    tecnico = request.args.get(
        "tecnico"
    )

    if tecnico:

        chamados_filtrados = []

        for c in chamados:

            tecnicos_db = db.session.execute(

                text("""

                    SELECT tecnico

                    FROM chamados_tecnicos

                    WHERE chamado_id = :id

                """),

                {
                    "id": c.id
                }

            ).fetchall()

            tecnicos = [

                t.tecnico

                for t in tecnicos_db

            ]

            if tecnico in tecnicos:

                chamados_filtrados.append(c)

        chamados = chamados_filtrados

    arquivo = "relatorio_chamados.pdf"

    doc = SimpleDocTemplate(
        arquivo
    )

    styles = getSampleStyleSheet()

    elementos = []

    elementos.append(

        Paragraph(

            "Relatório de Chamados",

            styles["Title"]

        )

    )

    elementos.append(
        Spacer(1, 20)
    )

    for c in chamados:

        tecnicos_db = db.session.execute(

            text("""

                SELECT tecnico

                FROM chamados_tecnicos

                WHERE chamado_id = :id

            """),

            {
                "id": c.id
            }

        ).fetchall()

        tecnicos = [

            t.tecnico

            for t in tecnicos_db

        ]

        texto = f"""

        <b>ID:</b> {c.id}<br/>

        <b>Colaborador:</b> {c.usuario}<br/>

        <b>Setor:</b> {c.setor}<br/>

        <b>Máquina:</b> {c.maquina}<br/>

        <b>Status:</b> {c.status}<br/>

        <b>Categoria:</b> {c.categoria}<br/>

        <b>Técnicos:</b> {', '.join(tecnicos)}<br/>

        <b>Solução:</b> {c.solucao or ''}<br/><br/>

        """

        elementos.append(

            Paragraph(

                texto,

                styles["BodyText"]

            )

        )

        elementos.append(
            Spacer(1, 12)
        )

    doc.build(elementos)

    return send_file(
        arquivo,
        as_attachment=True
    )
# =========================
# EXPORTAR XLSX
# =========================

@app.route("/exportar/xlsx")
def exportar_xlsx():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    chamados = db.session.execute(

        text("""

            SELECT *

            FROM chamados

            ORDER BY id DESC

        """)

    ).fetchall()

    tecnico = request.args.get(
        "tecnico"
    )

    if tecnico:

        chamados_filtrados = []

        for c in chamados:

            tecnicos_db = db.session.execute(

                text("""

                    SELECT tecnico

                    FROM chamados_tecnicos

                    WHERE chamado_id = :id

                """),

                {
                    "id": c.id
                }

            ).fetchall()

            tecnicos = [

                t.tecnico

                for t in tecnicos_db

            ]

            if tecnico in tecnicos:

                chamados_filtrados.append(c)

        chamados = chamados_filtrados

    wb = Workbook()

    ws = wb.active

    ws.title = "Chamados"

    headers = [

        "ID",
        "Colaborador",
        "Setor",
        "Máquina",
        "Status",
        "Categoria",
        "Técnicos",
        "Solução"

    ]

    ws.append(headers)

    for c in chamados:

        tecnicos_db = db.session.execute(

            text("""

                SELECT tecnico

                FROM chamados_tecnicos

                WHERE chamado_id = :id

            """),

            {
                "id": c.id
            }

        ).fetchall()

        tecnicos = [

            t.tecnico

            for t in tecnicos_db

        ]

        ws.append([

            c.id,
            c.usuario,
            c.setor,
            c.maquina,
            c.status,
            c.categoria,
            ", ".join(tecnicos),
            c.solucao or ""

        ])

    arquivo = "relatorio_chamados.xlsx"

    wb.save(arquivo)

    return send_file(
        arquivo,
        as_attachment=True
    )
# =========================
# EDITAR CHAMADO
# =========================

@app.route("/editar_chamado/<int:id>", methods=["POST"])
def editar_chamado(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    db.session.execute(
        db.text("""
            UPDATE chamados

            SET
                usuario = :usuario,
                setor = :setor,
                maquina = :maquina,
                categoria = :categoria,
                solucao = :solucao

            WHERE id = :id
        """),
        {
            "id": id,
            "usuario": request.form.get("usuario", "Não identificado"),
            "setor": request.form.get("setor", "Pendente"),
            "maquina": request.form.get("maquina", ""),
            "categoria": request.form.get("categoria", ""),
            "solucao": request.form.get("solucao", "")
        }
    )

    db.session.commit()

    flash(
        "Chamado atualizado!",
        "success"
    )

    return redirect(f"/chamado/{id}")
# =========================
# REABRIR CHAMADO
# =========================

@app.route("/reabrir/<int:id>")
def reabrir(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    db.session.execute(

        db.text("""

            UPDATE chamados

            SET

                status = 'em_andamento',

                finalizado_em = NULL

            WHERE id = :id

        """),

        {
            "id": id
        }

    )

    db.session.commit()

    flash(
        "Chamado reaberto!",
        "success"
    )

    return redirect("/chamados")
# =========================
# EXECUTAR
# =========================
# =========================
# INVENTÁRIO
# =========================

@app.route("/inventario")
def inventario():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:
        return redirect("/")

    ativos = db.session.execute(

        db.text("""

            SELECT *

            FROM ativos

            ORDER BY id DESC

        """)

    ).fetchall()

    total_ativos = db.session.execute(

        db.text("""

            SELECT COUNT(*)

            FROM ativos

        """)

    ).scalar()

    total_setores = db.session.execute(

        db.text("""

            SELECT COUNT(DISTINCT setor)

            FROM ativos

        """)

    ).scalar()

    total_colaboradores = db.session.execute(

        db.text("""

            SELECT COUNT(DISTINCT usuario_atual)

            FROM ativos

            WHERE usuario_atual IS NOT NULL

        """)

    ).scalar()

    maquinas_chamado = db.session.execute(

        db.text("""

            SELECT COUNT(DISTINCT maquina)

            FROM chamados

            WHERE status != 'Finalizado'

        """)

    ).scalar()

    return render_template(

        "inventario.html",

        usuario=usuario,

        ativos=ativos,

        total_ativos=total_ativos,

        total_setores=total_setores,

        total_colaboradores=total_colaboradores,

        maquinas_chamado=maquinas_chamado

    )
# =========================
# EXPORTAR INVENTÁRIO XLSX
# =========================

@app.route("/exportar_inventario_xlsx")
def exportar_inventario_xlsx():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:
        return redirect("/")

    ativos = db.session.execute(

        db.text("""

            SELECT

                a.id,
                a.id_maquina,
                a.marca,
                a.modelo,
                a.sistema_operacional,
                a.memoria_ram,
                a.armazenamento,
                a.usuario_atual,
                a.setor,

                CASE

                    WHEN EXISTS (

                        SELECT 1

                        FROM chamados c

                        WHERE c.maquina = a.id_maquina

                        AND LOWER(c.status) NOT IN (
                            'finalizado',
                            'resolvido'
                        )

                        AND COALESCE(c.arquivado, false) = false

                    )

                    THEN 'Sim'

                    ELSE 'Não'

                END AS em_chamado

            FROM ativos a

            ORDER BY a.setor, a.usuario_atual, a.id_maquina

        """)

    ).fetchall()

    wb = Workbook()

    ws = wb.active

    ws.title = "Inventário"

    # =========================
    # ESTILOS
    # =========================

    azul_escuro = "0F172A"
    azul = "1D4ED8"
    verde = "22C55E"
    vermelho = "EF4444"
    branco = "FFFFFF"
    cinza_claro = "E2E8F0"
    cinza_texto = "475569"

    fill_titulo = PatternFill(
        "solid",
        fgColor=azul_escuro
    )

    fill_header = PatternFill(
        "solid",
        fgColor=azul
    )

    fill_ok = PatternFill(
        "solid",
        fgColor="DCFCE7"
    )

    fill_alerta = PatternFill(
        "solid",
        fgColor="FEE2E2"
    )

    fonte_titulo = Font(
        color=branco,
        bold=True,
        size=16
    )

    fonte_header = Font(
        color=branco,
        bold=True
    )

    fonte_kpi = Font(
        bold=True,
        size=14,
        color=azul_escuro
    )

    borda = Border(
        left=Side(style="thin", color=cinza_claro),
        right=Side(style="thin", color=cinza_claro),
        top=Side(style="thin", color=cinza_claro),
        bottom=Side(style="thin", color=cinza_claro)
    )

    alinhamento_centro = Alignment(
        horizontal="center",
        vertical="center"
    )

    alinhamento_texto = Alignment(
        vertical="center",
        wrap_text=True
    )

    # =========================
    # TÍTULO
    # =========================

    ws.merge_cells("A1:J1")

    ws["A1"] = "Inventário de Máquinas - Vyron ITSM"

    ws["A1"].fill = fill_titulo
    ws["A1"].font = fonte_titulo
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[1].height = 30

    # =========================
    # RESUMO
    # =========================

    total_ativos = len(ativos)

    setores = set()

    colaboradores = set()

    maquinas_chamado = 0

    for a in ativos:

        if a.setor:
            setores.add(a.setor)

        if a.usuario_atual:
            colaboradores.add(a.usuario_atual)

        if a.em_chamado == "Sim":
            maquinas_chamado += 1

    resumo = [
        ["Total de ativos", total_ativos],
        ["Setores", len(setores)],
        ["Colaboradores", len(colaboradores)],
        ["Máquinas em chamado", maquinas_chamado]
    ]

    linha_resumo = 3

    for label, valor in resumo:

        ws.cell(
            row=linha_resumo,
            column=1,
            value=label
        )

        ws.cell(
            row=linha_resumo,
            column=2,
            value=valor
        )

        ws.cell(
            row=linha_resumo,
            column=1
        ).font = Font(
            bold=True,
            color=cinza_texto
        )

        ws.cell(
            row=linha_resumo,
            column=2
        ).font = fonte_kpi

        linha_resumo += 1

    # =========================
    # CABEÇALHO DA TABELA
    # =========================

    headers = [

        "ID Interno",
        "ID Máquina",
        "Marca",
        "Modelo",
        "Sistema",
        "Memória RAM",
        "Armazenamento",
        "Usuário",
        "Setor",
        "Chamado Aberto"

    ]

    linha_header = 9

    for col, header in enumerate(headers, start=1):

        celula = ws.cell(
            row=linha_header,
            column=col,
            value=header
        )

        celula.fill = fill_header
        celula.font = fonte_header
        celula.border = borda
        celula.alignment = alinhamento_centro

    # =========================
    # DADOS
    # =========================

    linha = linha_header + 1

    for a in ativos:

        dados = [

            a.id,
            a.id_maquina,
            a.marca,
            a.modelo,
            a.sistema_operacional,
            a.memoria_ram,
            a.armazenamento,
            a.usuario_atual,
            a.setor,
            a.em_chamado

        ]

        for col, valor in enumerate(dados, start=1):

            celula = ws.cell(
                row=linha,
                column=col,
                value=valor
            )

            celula.border = borda
            celula.alignment = alinhamento_texto

            if col == 10:

                celula.alignment = alinhamento_centro

                if valor == "Sim":
                    celula.fill = fill_alerta
                    celula.font = Font(
                        bold=True,
                        color=vermelho
                    )
                else:
                    celula.fill = fill_ok
                    celula.font = Font(
                        bold=True,
                        color="15803D"
                    )

        linha += 1

    # =========================
    # FILTROS E CONGELAMENTO
    # =========================

    ultima_linha = linha - 1

    ws.auto_filter.ref = f"A{linha_header}:J{ultima_linha}"

    ws.freeze_panes = "A10"

    # =========================
    # LARGURAS
    # =========================

    larguras = {

        "A": 12,
        "B": 14,
        "C": 18,
        "D": 28,
        "E": 18,
        "F": 16,
        "G": 20,
        "H": 28,
        "I": 18,
        "J": 18

    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[coluna].width = largura

    # =========================
    # ABA RESUMO POR SETOR
    # =========================

    ws2 = wb.create_sheet("Resumo por Setor")

    ws2.merge_cells("A1:C1")

    ws2["A1"] = "Resumo por Setor"

    ws2["A1"].fill = fill_titulo
    ws2["A1"].font = fonte_titulo
    ws2["A1"].alignment = alinhamento_centro

    ws2.append([])

    ws2.append([
        "Setor",
        "Total de Máquinas",
        "Máquinas com Chamado"
    ])

    for cell in ws2[3]:

        cell.fill = fill_header
        cell.font = fonte_header
        cell.border = borda
        cell.alignment = alinhamento_centro

    resumo_setores = {}

    for a in ativos:

        setor = a.setor or "Não informado"

        if setor not in resumo_setores:

            resumo_setores[setor] = {
                "total": 0,
                "chamados": 0
            }

        resumo_setores[setor]["total"] += 1

        if a.em_chamado == "Sim":

            resumo_setores[setor]["chamados"] += 1

    linha_setor = 4

    for setor, dados in sorted(resumo_setores.items()):

        ws2.cell(
            row=linha_setor,
            column=1,
            value=setor
        )

        ws2.cell(
            row=linha_setor,
            column=2,
            value=dados["total"]
        )

        ws2.cell(
            row=linha_setor,
            column=3,
            value=dados["chamados"]
        )

        for col in range(1, 4):

            ws2.cell(
                row=linha_setor,
                column=col
            ).border = borda

            ws2.cell(
                row=linha_setor,
                column=col
            ).alignment = alinhamento_centro

        linha_setor += 1

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 24

    # =========================
    # EXPORTAÇÃO EM MEMÓRIA
    # =========================

    arquivo = BytesIO()

    wb.save(arquivo)

    arquivo.seek(0)

    return send_file(

        arquivo,

        as_attachment=True,

        download_name="inventario_maquinas_vyron.xlsx",

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

# =========================
# EXPORTAR INVENTÁRIO RH XLSX
# =========================

@app.route("/exportar_inventario_rh_xlsx")
def exportar_inventario_rh_xlsx():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:
        return redirect("/")

    ativos = db.session.execute(

        db.text("""

            SELECT

                id_maquina,
                marca,
                modelo,
                sistema_operacional,
                memoria_ram,
                armazenamento,
                usuario_atual,
                setor

            FROM ativos

            ORDER BY setor, usuario_atual, id_maquina

        """)

    ).fetchall()

    wb = Workbook()

    ws = wb.active

    ws.title = "Inventário RH"

    ws.sheet_view.showGridLines = False

    # =========================
    # CORES
    # =========================

    azul_noite = "081426"
    azul_profundo = "0F2747"
    azul_vyron = "2563EB"
    azul_claro = "DBEAFE"
    verde = "22C55E"
    branco = "FFFFFF"
    cinza_borda = "CBD5E1"
    cinza_texto = "64748B"
    cinza_linha = "F8FAFC"
    cinza_linha_2 = "EEF2F7"

    # =========================
    # ESTILOS
    # =========================

    fill_titulo = PatternFill(
        "solid",
        fgColor=azul_noite
    )

    fill_subtitulo = PatternFill(
        "solid",
        fgColor=azul_profundo
    )

    fill_header = PatternFill(
        "solid",
        fgColor=azul_vyron
    )

    fill_kpi = PatternFill(
        "solid",
        fgColor=azul_claro
    )

    fill_linha = PatternFill(
        "solid",
        fgColor=cinza_linha
    )

    fill_linha_2 = PatternFill(
        "solid",
        fgColor=cinza_linha_2
    )

    fonte_titulo = Font(
        color=branco,
        bold=True,
        size=20
    )

    fonte_subtitulo = Font(
        color=branco,
        bold=True,
        size=11
    )

    fonte_header = Font(
        color=branco,
        bold=True,
        size=11
    )

    fonte_kpi_label = Font(
        color=cinza_texto,
        bold=True,
        size=10
    )

    fonte_kpi_valor = Font(
        color=azul_noite,
        bold=True,
        size=18
    )

    fonte_padrao = Font(
        color="0F172A",
        size=10
    )

    fonte_setor = Font(
        color=azul_vyron,
        bold=True,
        size=10
    )

    borda_fina = Border(
        left=Side(style="thin", color=cinza_borda),
        right=Side(style="thin", color=cinza_borda),
        top=Side(style="thin", color=cinza_borda),
        bottom=Side(style="thin", color=cinza_borda)
    )

    centro = Alignment(
        horizontal="center",
        vertical="center"
    )

    esquerda = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True
    )

    # =========================
    # CABEÇALHO PRINCIPAL
    # =========================

    ws.merge_cells("A1:H2")

    ws["A1"] = "Inventário de Máquinas - RH"

    ws["A1"].fill = fill_titulo
    ws["A1"].font = fonte_titulo
    ws["A1"].alignment = centro

    for row in range(1, 3):

        for col in range(1, 9):

            ws.cell(row=row, column=col).fill = fill_titulo

    ws.merge_cells("A3:H3")

    ws["A3"] = "Relatório gerado pelo Vyron ITSM • Controle interno de máquinas por usuário e setor"

    ws["A3"].fill = fill_subtitulo
    ws["A3"].font = fonte_subtitulo
    ws["A3"].alignment = centro

    for col in range(1, 9):

        ws.cell(row=3, column=col).fill = fill_subtitulo

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 24

        # =========================
    # RESUMO
    # =========================

    total_ativos = len(ativos)

    total_setores = len(set([
        a.setor
        for a in ativos
        if a.setor
    ]))

    total_usuarios = len(set([
        a.usuario_atual
        for a in ativos
        if a.usuario_atual
    ]))

    sem_usuario = len([
        a for a in ativos
        if not a.usuario_atual
    ])

    resumo = [
        ("Total de máquinas", total_ativos),
        ("Setores", total_setores),
        ("Usuários vinculados", total_usuarios),
        ("Sem usuário", sem_usuario)
    ]

    blocos = [
        ("A5:B5", "A6:B6", "A5", "A6"),
        ("C5:D5", "C6:D6", "C5", "C6"),
        ("E5:F5", "E6:F6", "E5", "E6"),
        ("G5:H5", "G6:H6", "G5", "G6")
    ]

    for i, (label, valor) in enumerate(resumo):

        intervalo_label, intervalo_valor, cel_label, cel_valor = blocos[i]

        ws.merge_cells(intervalo_label)
        ws.merge_cells(intervalo_valor)

        ws[cel_label] = label
        ws[cel_label].font = fonte_kpi_label
        ws[cel_label].alignment = centro
        ws[cel_label].fill = fill_kpi
        ws[cel_label].border = borda_fina

        ws[cel_valor] = valor
        ws[cel_valor].font = fonte_kpi_valor
        ws[cel_valor].alignment = centro
        ws[cel_valor].fill = fill_kpi
        ws[cel_valor].border = borda_fina

        for intervalo in [
            intervalo_label,
            intervalo_valor
        ]:

            inicio, fim = intervalo.split(":")

            col_inicio = ws[inicio].column
            col_fim = ws[fim].column
            row_inicio = ws[inicio].row
            row_fim = ws[fim].row

            for row in range(row_inicio, row_fim + 1):

                for col in range(col_inicio, col_fim + 1):

                    celula = ws.cell(
                        row=row,
                        column=col
                    )

                    celula.fill = fill_kpi
                    celula.border = borda_fina
                    celula.alignment = centro

    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 30
    # =========================
    # CABEÇALHO DA TABELA
    # =========================

    linha_header = 9

    headers = [
        "ID Máquina",
        "Marca",
        "Modelo",
        "Sistema",
        "Memória RAM",
        "Armazenamento",
        "Usuário",
        "Setor"
    ]

    for col, header in enumerate(headers, start=1):

        celula = ws.cell(
            row=linha_header,
            column=col,
            value=header
        )

        celula.fill = fill_header
        celula.font = fonte_header
        celula.border = borda_fina
        celula.alignment = centro

    ws.row_dimensions[linha_header].height = 26

    # =========================
    # DADOS
    # =========================

    linha = linha_header + 1

    for a in ativos:

        dados = [
            a.id_maquina,
            a.marca,
            a.modelo,
            a.sistema_operacional,
            a.memoria_ram,
            a.armazenamento,
            a.usuario_atual,
            a.setor
        ]

        fill_atual = fill_linha if linha % 2 == 0 else fill_linha_2

        for col, valor in enumerate(dados, start=1):

            celula = ws.cell(
                row=linha,
                column=col,
                value=valor or "-"
            )

            celula.fill = fill_atual
            celula.border = borda_fina
            celula.font = fonte_padrao
            celula.alignment = esquerda

            if col in [1, 4, 5, 6, 8]:

                celula.alignment = centro

            if col == 8:

                celula.font = fonte_setor

        ws.row_dimensions[linha].height = 24

        linha += 1

    # =========================
    # FILTRO, CONGELAMENTO E LARGURA
    # =========================

    ultima_linha = linha - 1

    ws.auto_filter.ref = f"A{linha_header}:H{ultima_linha}"

    ws.freeze_panes = "A10"

    larguras = {
        "A": 14,
        "B": 18,
        "C": 28,
        "D": 18,
        "E": 16,
        "F": 20,
        "G": 30,
        "H": 18
    }

    for coluna, largura in larguras.items():

        ws.column_dimensions[coluna].width = largura

    # =========================
    # RODAPÉ
    # =========================

    linha_rodape = ultima_linha + 3

    ws.merge_cells(
        start_row=linha_rodape,
        start_column=1,
        end_row=linha_rodape,
        end_column=8
    )

    ws.cell(
        row=linha_rodape,
        column=1,
        value="Relatório gerado automaticamente pelo Vyron ITSM."
    )

    ws.cell(
        row=linha_rodape,
        column=1
    ).font = Font(
        color=cinza_texto,
        italic=True,
        size=9
    )

    ws.cell(
        row=linha_rodape,
        column=1
    ).alignment = centro

    # =========================
    # CONFIGURAÇÃO DE IMPRESSÃO
    # =========================

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.print_title_rows = f"{linha_header}:{linha_header}"

    # =========================
    # EXPORTAÇÃO
    # =========================

    arquivo = BytesIO()

    wb.save(arquivo)

    arquivo.seek(0)

    return send_file(

        arquivo,

        as_attachment=True,

        download_name="inventario_rh_vyron.xlsx",

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )
# =========================
# USUÁRIOS
# =========================

@app.route("/usuarios")
def usuarios():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:
        return redirect("/")

    busca = request.args.get(
        "busca",
        ""
    ).strip()

    if busca:

        usuarios_db = db.session.execute(

            db.text("""

                SELECT *

                FROM usuarios

                WHERE

                    CAST(id AS TEXT)
                    LIKE :busca

                    OR

                    LOWER(nome)
                    LIKE LOWER(:busca)

                ORDER BY id ASC

            """),

            {
                "busca": f"%{busca}%"
            }

        ).fetchall()

    else:

        usuarios_db = db.session.execute(

            db.text("""

                SELECT *

                FROM usuarios

                ORDER BY id ASC

            """)

        ).fetchall()

    return render_template(

        "usuarios.html",

        usuarios=usuarios_db,

        usuario=usuario,

        busca=busca

    )
# =========================
# NOVO USUÁRIO
# =========================

@app.route(
    "/novo_usuario",
    methods=["GET", "POST"]
)
def novo_usuario():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    if request.method == "POST":

        dados = {

            "id": request.form.get("id"),

            "nome": request.form.get("nome"),

            "senha": generate_password_hash(
                request.form.get("senha")
            ),

            "tipo": request.form.get("tipo"),

            "setor": request.form.get("setor"),

            "email": request.form.get("email"),

            "ativo": True

        }

        existe = db.session.execute(

            db.text("""

                SELECT id

                FROM usuarios

                WHERE id = :id

                OR nome = :nome

            """),

            {
                "id": dados["id"],
                "nome": dados["nome"]
            }

        ).fetchone()

        if existe:

            flash(
                "Usuário já existe.",
                "danger"
            )

            return redirect("/novo_usuario")

        db.session.execute(

            db.text("""

                INSERT INTO usuarios (

                    id,
                    nome,
                    senha,
                    tipo,
                    setor,
                    email,
                    ativo

                )

                VALUES (

                    :id,
                    :nome,
                    :senha,
                    :tipo,
                    :setor,
                    :email,
                    :ativo

                )

            """),

            dados

        )

        db.session.commit()

        flash(
            "Usuário criado com sucesso!",
            "success"
        )

        return redirect("/usuarios")

    return render_template(
        "novo_usuario.html",
        usuario=usuario
    )

# =========================
# IMPORTAR USUÁRIOS XLSX
# =========================

@app.route(
    "/importar_usuarios",
    methods=["POST"]
)
def importar_usuarios():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    arquivo = request.files.get("arquivo")

    if not arquivo:

        flash(
            "Nenhum arquivo enviado.",
            "error"
        )

        return redirect("/usuarios")

    try:

        df = pd.read_excel(arquivo)

        for _, linha in df.iterrows():

            existe = db.session.execute(

                db.text("""

                    SELECT id

                    FROM usuarios

                    WHERE id = :id

                """),

                {
                    "id": int(linha["id"])
                }

            ).fetchone()

            if existe:
                continue

            senha_hash = generate_password_hash(
                str(linha["senha"])
            )

            db.session.execute(

                db.text("""

                    INSERT INTO usuarios (

                        id,
                        nome,
                        senha,
                        tipo,
                        setor,
                        email,
                        ativo

                    )

                    VALUES (

                        :id,
                        :nome,
                        :senha,
                        :tipo,
                        :setor,
                        :email,
                        true

                    )

                """),

                {

                    "id": int(linha["id"]),

                    "nome": str(linha["nome"]),

                    "senha": senha_hash,

                    "tipo": str(linha["tipo"]),

                    "setor": str(linha["setor"]),

                    "email": str(linha["email"])

                }

            )

        db.session.commit()

        flash(
            "Usuários importados com sucesso!",
            "success"
        )

    except Exception as e:

        flash(
            f"Erro ao importar planilha: {e}",
            "error"
        )

    return redirect("/usuarios")
# =========================
# EDITAR USUÁRIO
# =========================

@app.route(
    "/editar_usuario/<int:id>",
    methods=["GET", "POST"]
)
def editar_usuario(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    usuario_db = db.session.execute(

        db.text("""

            SELECT *

            FROM usuarios

            WHERE id = :id

        """),

        {
            "id": id
        }

    ).fetchone()

    if not usuario_db:
        return redirect("/usuarios")

    if request.method == "POST":

        nova_senha = request.form.get("senha")

        if nova_senha:

            db.session.execute(

                db.text("""

                    UPDATE usuarios

                    SET

                        nome = :nome,
                        senha = :senha,
                        tipo = :tipo,
                        setor = :setor,
                        email = :email,
                        ativo = :ativo

                    WHERE id = :id

                """),

                {

                    "id": id,

                    "nome": request.form.get("nome"),

                    "senha": generate_password_hash(
                        nova_senha
                    ),

                    "tipo": request.form.get("tipo"),

                    "setor": request.form.get("setor"),

                    "email": request.form.get("email"),

                    "ativo":
                    True if request.form.get("ativo")
                    == "true"
                    else False

                }

            )

        else:

            db.session.execute(

                db.text("""

                    UPDATE usuarios

                    SET

                        nome = :nome,
                        tipo = :tipo,
                        setor = :setor,
                        email = :email,
                        ativo = :ativo

                    WHERE id = :id

                """),

                {

                    "id": id,

                    "nome": request.form.get("nome"),

                    "tipo": request.form.get("tipo"),

                    "setor": request.form.get("setor"),

                    "email": request.form.get("email"),

                    "ativo":
                    True if request.form.get("ativo")
                    == "true"
                    else False

                }

            )

        db.session.commit()

        flash(
            "Usuário atualizado com sucesso!",
            "success"
        )

        return redirect("/usuarios")

    return render_template(

        "editar_usuario.html",

        usuario_db=usuario_db,

        usuario=usuario
    )
# =========================
# DESATIVAR USUÁRIO
# =========================

@app.route("/desativar_usuario/<int:id>")
def desativar_usuario(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    db.session.execute(

        db.text("""

            UPDATE usuarios

            SET ativo = false

            WHERE id = :id

        """),

        {
            "id": id
        }

    )

    db.session.commit()

    flash(
        "Usuário desativado.",
        "warning"
    )

    return redirect("/usuarios")

# =========================
# RESETAR SENHA
# =========================

@app.route("/resetar_senha/<int:id>")
def resetar_senha(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:
        return redirect("/")

    nova_senha = "123456"

    senha_hash = generate_password_hash(
        nova_senha
    )

    db.session.execute(

        db.text("""

            UPDATE usuarios

            SET senha = :senha,

            precisa_trocar_senha = TRUE

            WHERE id = :id

        """),

        {
            "senha": senha_hash,
            "id": id
        }

    )

    db.session.commit()

    flash(
        "Senha resetada para: 123456",
        "success"
    )

    return redirect("/usuarios")
# =========================
# IMPORTAR INVENTÁRIOS
# =========================

@app.route("/importar_inventario")
def importar_inventario():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:
        return redirect("/")

    pasta = "Planilhas"

    arquivos = [

        arq

        for arq in os.listdir(pasta)

        if arq.endswith(".xlsx")

    ]

    total_importados = 0

    for arquivo in arquivos:

        caminho = os.path.join(
            pasta,
            arquivo
        )

        df = pd.read_excel(caminho)

        df.columns = [

            str(col).strip().upper()

            for col in df.columns

        ]

        for _, row in df.iterrows():

            id_maquina = str(
                row.get("ID", "")
            )

            existe = db.session.execute(

                db.text("""

                    SELECT id

                    FROM ativos

                    WHERE id_maquina = :id_maquina

                """),

                {
                    "id_maquina": id_maquina
                }

            ).fetchone()

            dados = {

                "id_maquina": id_maquina,

                "marca": str(
                    row["MARCA"]
                ),

                "modelo": str(
                    row["MODELO"]
                ),

                "sistema_operacional": str(
                    row["SISTEMA OPERACIONAL"]
                ),

                "memoria_ram": str(
                    row["MEMORIA RAM"]
                ),

                "armazenamento": str(
                    row["ARMAZENAMENTO"]
                ),

                "usuario_atual": str(
                    row["USUARIO"]
                ),

                "setor": str(
                    row["SETOR"]
                )

            }

            if existe:

                db.session.execute(

                    db.text("""

                        UPDATE ativos

                        SET

                            marca = :marca,

                            modelo = :modelo,

                            sistema_operacional =
                            :sistema_operacional,

                            memoria_ram =
                            :memoria_ram,

                            armazenamento =
                            :armazenamento,

                            usuario_atual =
                            :usuario_atual,

                            setor = :setor

                        WHERE id_maquina =
                        :id_maquina

                    """),

                    dados

                )

            else:

                db.session.execute(

                    db.text("""

                        INSERT INTO ativos (

                            id_maquina,
                            marca,
                            modelo,
                            sistema_operacional,
                            memoria_ram,
                            armazenamento,
                            usuario_atual,
                            setor

                        )

                        VALUES (

                            :id_maquina,
                            :marca,
                            :modelo,
                            :sistema_operacional,
                            :memoria_ram,
                            :armazenamento,
                            :usuario_atual,
                            :setor

                        )

                    """),

                    dados

                )

            total_importados += 1

    db.session.commit()

    return f"""

    Inventários importados com sucesso!

    Total de registros processados:
    {total_importados}

    """

# =========================
# PESQUISA INVENTÁRIO
# =========================

@app.route(
    "/pesquisar_ativos",
    methods=["GET"]
)
def pesquisar_ativos():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:
        return redirect("/")

    termo = request.args.get(
        "termo",
        ""
    ).strip()

    ram = request.args.get(
        "ram",
        ""
    ).strip()

    armazenamento = request.args.get(
        "armazenamento",
        ""
    ).strip()

    sistema = request.args.get(
        "sistema",
        ""
    ).strip()

    marca = request.args.get(
        "marca",
        ""
    ).strip()

    setores_fixos = [
        "backoffice",
        "hunter",
        "bl",
        "farmer",
        "pós",
        "pos",
        "ti",
        "admin"
    ]

    query = """

        SELECT

            id,
            id_maquina,
            marca,
            modelo,
            sistema_operacional,
            memoria_ram,
            armazenamento,
            usuario_atual,
            setor

        FROM ativos

        WHERE 1=1

    """

    params = {}

    if termo:

        if termo.lower() in setores_fixos:

            query += """

                AND LOWER(setor) = LOWER(:termo)

            """

            params["termo"] = termo

        else:

            query += """

                AND (

                    LOWER(id_maquina)
                    LIKE LOWER(:termo)

                    OR

                    LOWER(usuario_atual)
                    LIKE LOWER(:termo)

                    OR

                    LOWER(setor)
                    LIKE LOWER(:termo)

                    OR

                    LOWER(modelo)
                    LIKE LOWER(:termo)

                    OR

                    LOWER(marca)
                    LIKE LOWER(:termo)

                )

            """

            params["termo"] = f"%{termo}%"

    if ram:

        query += """

            AND LOWER(memoria_ram)
            LIKE LOWER(:ram)

        """

        params["ram"] = f"%{ram}%"

    if armazenamento:

        query += """

            AND LOWER(armazenamento)
            LIKE LOWER(:armazenamento)

        """

        params["armazenamento"] = f"%{armazenamento}%"

    if sistema:

        query += """

            AND LOWER(sistema_operacional)
            LIKE LOWER(:sistema)

        """

        params["sistema"] = f"%{sistema}%"

    if marca:

        query += """

            AND LOWER(marca)
            LIKE LOWER(:marca)

        """

        params["marca"] = f"%{marca}%"

    query += """

        ORDER BY id DESC

    """

    ativos = db.session.execute(
        db.text(query),
        params
    ).fetchall()

    return render_template(

        "pesquisa_ativos.html",

        ativos=ativos,

        termo=termo,

        ram=ram,

        armazenamento=armazenamento,

        sistema=sistema,

        marca=marca,

        usuario=usuario

    )
@app.route(
    "/chamado_rapido",
    methods=["GET", "POST"]
)
def chamado_rapido():

    if request.method == "POST":

        descricao = request.form.get("descricao", "")
        maquina = request.form.get("maquina", "")
        local = request.form.get("local", "")

        db.session.execute(
            db.text("""
                INSERT INTO chamados (
                    usuario,
                    descricao,
                    setor,
                    maquina,
                    status,
                    categoria,
                    solucao,
                    origem,
                    local_informado
                )
                VALUES (
                    'Não identificado',
                    :descricao,
                    'Pendente',
                    :maquina,
                    'aberto',
                    '',
                    '',
                    'QR Code',
                    :local
                )
            """),
            {
                "descricao": descricao,
                "maquina": maquina,
                "local": local
            }
        )

        chamado_criado = db.session.execute(
            db.text("""
                SELECT id
                FROM chamados
                WHERE origem = 'QR Code'
                ORDER BY id DESC
                LIMIT 1
            """)
        ).fetchone()

        if chamado_criado:

            tecnicos = db.session.execute(
                db.text("""
                    SELECT id
                    FROM usuarios
                    WHERE tipo IN ('ti')
                    AND ativo = true
                """)
            ).fetchall()

            for tecnico in tecnicos:

                criar_notificacao(
                    usuario_id=tecnico.id,
                    titulo="Chamado rápido aberto",
                    mensagem=f"Um chamado via QR Code foi aberto. #{chamado_criado.id}",
                    tipo="chamado",
                    link=f"/chamado/{chamado_criado.id}"
                )

        db.session.commit()

        return render_template("chamado_rapido_sucesso.html")

    return render_template("chamado_rapido.html")
# =========================
# EDITAR POSIÇÃO
# =========================

@app.route(
    "/editar_posicao/<path:posicao>",
    methods=["GET", "POST"]
)
def editar_posicao(posicao):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:
        return redirect("/")

    posicao = urllib.parse.unquote(posicao)

    posicao_db = db.session.execute(

        db.text("""

            SELECT *

            FROM mapa_posicoes

            WHERE posicao = :posicao

        """),

        {
            "posicao": posicao
        }

    ).fetchone()

    if not posicao_db:

        flash(
            "Posição não encontrada.",
            "error"
        )

        return redirect("/")

    ativo = None

    if posicao_db.maquina:

        ativo = db.session.execute(

            db.text("""

                SELECT *

                FROM ativos

                WHERE id_maquina = :maquina

            """),

            {
                "maquina": posicao_db.maquina
            }

        ).fetchone()

    if request.method == "POST":

        nova_posicao = request.form.get(
            "nova_posicao"
        )

        nova_maquina = request.form.get(
            "maquina"
        )

        colaborador = request.form.get(
            "colaborador"
        )

        # =========================
        # ATUALIZA MAPA
        # =========================

        db.session.execute(

            db.text("""

                UPDATE mapa_posicoes

                SET

                    posicao = :nova_posicao,
                    maquina = :maquina

                WHERE id = :id

            """),

            {
                "nova_posicao": nova_posicao,
                "maquina": nova_maquina,
                "id": posicao_db.id
            }

        )

        # =========================
        # ATUALIZA ATIVO
        # =========================

        if nova_maquina:

            db.session.execute(

                db.text("""

                    UPDATE ativos

                    SET usuario_atual = :usuario

                    WHERE id_maquina = :maquina

                """),

                {
                    "usuario": colaborador,
                    "maquina": nova_maquina
                }

            )

        db.session.commit()

        flash(
            "Posição atualizada com sucesso!",
            "success"
        )

        if posicao_db.sala == "BL":
            return redirect("/mapa_bl")

        return redirect("/mapa_hunter")

    return render_template(

        "editar_posicao.html",

        usuario=usuario,

        posicao=posicao_db,

        ativo=ativo

    )
# =========================
# EXCLUIR POSIÇÃO
# =========================

@app.route("/excluir_posicao/<path:posicao>")
def excluir_posicao(posicao):

    usuario = session.get("usuario")

    if not usuario:

        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:

        return redirect("/")

    sala_db = db.session.execute(

        db.text("""

            SELECT sala

            FROM mapa_posicoes

            WHERE posicao = :posicao

        """),

        {
            "posicao": posicao
        }

    ).fetchone()

    db.session.execute(

        db.text("""

            DELETE FROM mapa_posicoes

            WHERE posicao = :posicao

        """),

        {
            "posicao": posicao
        }

    )

    db.session.commit()

    flash(
        "Posição excluída com sucesso!",
        "success"
    )

    if sala_db:

        if sala_db.sala == "BL":

            return redirect(
                url_for("mapa_bl")
            )

        return redirect(
            url_for("mapa_hunter")
        )

    return redirect("/")
# =========================
# EDITAR ATIVO
# =========================

@app.route(
    "/editar_ativo/<int:id>",
    methods=["GET", "POST"]
)
def editar_ativo(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:
        return redirect("/")

    ativo = db.session.execute(

        db.text("""

            SELECT *

            FROM ativos

            WHERE id = :id

        """),

        {
            "id": id
        }

    ).fetchone()

    if request.method == "POST":

        dados = {

            "id": id,

            "marca": request.form.get("marca"),

            "modelo": request.form.get("modelo"),

            "sistema_operacional":
            request.form.get(
                "sistema_operacional"
            ),

            "memoria_ram":
            request.form.get(
                "memoria_ram"
            ),

            "armazenamento":
            request.form.get(
                "armazenamento"
            ),

            "usuario_atual":
            request.form.get(
                "usuario_atual"
            ),

            "setor":
            request.form.get(
                "setor"
            )

        }

        db.session.execute(

            db.text("""

                UPDATE ativos

                SET

                    marca = :marca,

                    modelo = :modelo,

                    sistema_operacional =
                    :sistema_operacional,

                    memoria_ram =
                    :memoria_ram,

                    armazenamento =
                    :armazenamento,

                    usuario_atual =
                    :usuario_atual,

                    setor = :setor

                WHERE id = :id

            """),

            dados

        )

        db.session.commit()

        return redirect(
            "/pesquisar_ativos"
        )

    return render_template(

        "editar_ativo.html",

        ativo=ativo,

        usuario=usuario
    )

# =========================
# MAPA HUNTER
# =========================

@app.route("/mapa_hunter")
def mapa_hunter():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:
        return redirect("/")

    # =========================
    # ATIVOS HUNTER
    # =========================

    maquinas = db.session.execute(

        db.text("""

            SELECT
                id,
                id_maquina,
                usuario_atual,
                marca,
                modelo,
                setor

            FROM ativos

            WHERE LOWER(setor)
            LIKE LOWER('%hunter%')

        """)

    ).fetchall()

    # =========================
    # CHAMADOS ABERTOS
    # =========================

    chamados_abertos = db.session.execute(

        db.text("""

            SELECT DISTINCT maquina

            FROM chamados

            WHERE status != 'Finalizado'

        """)

    ).fetchall()

    maquinas_com_chamado = [

        c.maquina

        for c in chamados_abertos

        if c.maquina

    ]

    # =========================
    # POSIÇÕES HUNTER
    # =========================

    mapa_db = db.session.execute(

        db.text("""

            SELECT
                id,
                posicao,
                maquina,
                colaborador,
                sala

            FROM mapa_posicoes

            WHERE sala = 'Hunter'

            ORDER BY posicao

        """)

    ).fetchall()

    # =========================
    # MONTAGEM MAPA
    # =========================

    mapa = {}

    for item in mapa_db:

        status = "livre"

        if item.maquina in maquinas_com_chamado:

            status = "problema"

        maquina_info = None

        for maquina in maquinas:

            if (
                str(maquina.id_maquina).strip()
                ==
                str(item.maquina).strip()
            ):

                maquina_info = {

                    "modelo": maquina.modelo,
                    "usuario_atual": maquina.usuario_atual,
                    "marca": maquina.marca,
                    "setor": maquina.setor

                }

                break

        mapa[item.posicao] = {

            "maquina": item.maquina,

            "colaborador": (
                maquina_info["usuario_atual"]
                if maquina_info
                else item.colaborador
            ),

            "status": status,

            "info": maquina_info

        }

    # =========================
    # CONTADORES
    # =========================

    total_desktops = len([

        item for item in mapa_db

        if "NOTEBOOK" not in item.posicao.upper()

    ])

    total_notebooks = len([

        item for item in mapa_db

        if "NOTEBOOK" in item.posicao.upper()

    ])

    total_chamados = len(maquinas_com_chamado)

    # =========================
    # RENDER
    # =========================

    return render_template(

        "mapa_hunter.html",

        usuario=usuario,

        maquinas=maquinas,

        mapa=mapa,

        mapa_db=mapa_db,

        maquinas_com_chamado=maquinas_com_chamado,

        total_desktops=total_desktops,

        total_notebooks=total_notebooks,

        total_chamados=total_chamados

    )
# =========================
# MAPA BL
# =========================

@app.route("/mapa_bl")
def mapa_bl():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    maquinas = db.session.execute(

        db.text("""

            SELECT
                id_maquina,
                usuario_atual,
                marca,
                modelo,
                setor

            FROM ativos

        """)

    ).fetchall()

    chamados_abertos = db.session.execute(

        db.text("""

            SELECT DISTINCT maquina

            FROM chamados

            WHERE status != 'Finalizado'

        """)

    ).fetchall()

    maquinas_com_chamado = [

        c.maquina

        for c in chamados_abertos

        if c.maquina

    ]

    mapa_db = db.session.execute(

        db.text("""

            SELECT
                id,
                posicao,
                maquina,
                sala

            FROM mapa_posicoes

            WHERE sala = 'BL'

            ORDER BY posicao

        """)

    ).fetchall()

    mapa = {}

    for item in mapa_db:

        status = "livre"

        if item.maquina in maquinas_com_chamado:
            status = "problema"

        maquina_info = None

        for maquina in maquinas:

            if maquina.id_maquina == item.maquina:

                maquina_info = {

                    "modelo": maquina.modelo,
                    "usuario_atual": maquina.usuario_atual,
                    "marca": maquina.marca,
                    "setor": maquina.setor

                }

                break

        mapa[item.posicao] = {

            "maquina": item.maquina,

            "colaborador":
            maquina_info["usuario_atual"]
            if maquina_info
            else None,

            "status": status,

            "info": maquina_info

        }

    total_desktops = len([

        item for item in mapa_db

        if (
            item.posicao.startswith("1015")
            or
            item.posicao.startswith("1016")
        )

    ])

    total_notebooks = len([

        item for item in mapa_db

        if "NOTEBOOK" in item.posicao.upper()

    ])

    total_chamados = len(maquinas_com_chamado)

    return render_template(

        "mapa_bl.html",

        usuario=usuario,
        maquinas=maquinas,
        mapa=mapa,
        mapa_db=mapa_db,
        maquinas_com_chamado=maquinas_com_chamado,

        total_desktops=total_desktops,
        total_notebooks=total_notebooks,
        total_chamados=total_chamados

    )
# =========================
# NOVA POSIÇÃO
# =========================

@app.route(
    "/nova_posicao",
    methods=["GET", "POST"]
)
def nova_posicao():

    usuario = session.get("usuario")

    if not usuario:

        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:

        return redirect("/")

    if request.method == "POST":

        sala = request.form.get(
            "sala"
        )

        posicao = request.form.get(
            "posicao"
        )

        maquina = request.form.get(
            "maquina"
        )

        colaborador = request.form.get(
            "colaborador"
        )

        # limpa espaços
        if sala:
            sala = sala.strip()

        if posicao:
            posicao = posicao.strip()

        if maquina:
            maquina = maquina.strip()

        if colaborador:
            colaborador = colaborador.strip()

        # verifica se já existe
        existe = db.session.execute(

            db.text("""

                SELECT id

                FROM mapa_posicoes

                WHERE posicao = :posicao

            """),

            {
                "posicao": posicao
            }

        ).fetchone()

        if existe:

            flash(
                "Essa posição já existe.",
                "error"
            )

            return redirect("/nova_posicao")

        db.session.execute(

            db.text("""

                INSERT INTO mapa_posicoes (

                    sala,
                    posicao,
                    maquina,
                    colaborador

                )

                VALUES (

                    :sala,
                    :posicao,
                    :maquina,
                    :colaborador

                )

            """),

            {
                "sala": sala,
                "posicao": posicao,
                "maquina": maquina,
                "colaborador": colaborador
            }

        )

        db.session.commit()

        flash(
            "Posição criada com sucesso!",
            "success"
        )

        if sala == "BL":

            return redirect("/mapa_bl")

        return redirect("/mapa_hunter")

    return render_template(

        "nova_posicao.html",

        usuario=usuario

    )

# =========================
# MOVIMENTAR MÁQUINA
# =========================

@app.route(
    "/movimentar_maquina",
    methods=["GET", "POST"]
)
def movimentar_maquina():

    usuario = session.get("usuario")

    if not usuario:

        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:

        return redirect("/")

    posicoes = db.session.execute(

        db.text("""

            SELECT *

            FROM mapa_posicoes

            ORDER BY sala, posicao

        """)

    ).fetchall()

    if request.method == "POST":

        origem = request.form.get(
            "origem"
        )

        destino = request.form.get(
            "destino"
        )

        origem_db = db.session.execute(

            db.text("""

                SELECT *

                FROM mapa_posicoes

                WHERE posicao = :origem

            """),

            {
                "origem": origem
            }

        ).fetchone()

        destino_db = db.session.execute(

            db.text("""

                SELECT *

                FROM mapa_posicoes

                WHERE posicao = :destino

            """),

            {
                "destino": destino
            }

        ).fetchone()

        if not origem_db or not destino_db:

            flash(
                "Posição inválida.",
                "error"
            )

            return redirect(
                "/movimentar_maquina"
            )
        ativo_origem = None

        if origem_db.maquina:

            ativo_origem = db.session.execute(

                db.text("""

                    SELECT *

                    FROM ativos

                    WHERE id_maquina = :maquina

                """),

                {
                    "maquina": origem_db.maquina
                }

            ).fetchone()
        # move dados
        db.session.execute(

            db.text("""

                UPDATE mapa_posicoes

                SET

                    maquina = :maquina,

                    colaborador = :colaborador

                WHERE posicao = :destino

            """),

            {
                "maquina": origem_db.maquina,
                "colaborador": (
                    ativo_origem.usuario_atual
                    if ativo_origem
                    else None
                ),
                "destino": destino
            }

        )

        # limpa origem
        db.session.execute(

            db.text("""

                UPDATE mapa_posicoes

                SET

                    maquina = NULL,

                    colaborador = NULL

                WHERE posicao = :origem

            """),

            {
                "origem": origem
            }

        )
        
        #histórico
        db.session.execute(
            db.text("""
                    INSERT INTO movimentacoes_mapa(
                    maquina,
                    colaborador,
                    origem,
                    destino,
                    usuario_responsavel
                    )
                    
                    VALUES(
                    :maquina,
                    :colaborador,
                    :origem,
                    :destino,
                    :usuario
                    )
                    """),
                {
                    "maquina": origem_db.maquina,
                    "colaborador": origem_db.colaborador,
                    "origem": origem,
                    "destino": destino,
                    "usuario": usuario["nome"]
                }
        )

        db.session.commit()

        flash(
            "Movimentação realizada com sucesso!",
            "success"
        )

        if origem_db.sala == "BL":

            return redirect("/mapa_bl")

        return redirect("/mapa_hunter")

    return render_template(

        "movimentar_maquina.html",

        usuario=usuario,

        posicoes=posicoes

    )
# =========================
# NOVO ATIVO
# =========================

@app.route(
    "/novo_ativo",
    methods=["GET", "POST"]
)
def novo_ativo():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    if request.method == "POST":

        db.session.execute(

            db.text("""

                INSERT INTO ativos (

                    id_maquina,
                    marca,
                    modelo,
                    sistema_operacional,
                    memoria_ram,
                    armazenamento,
                    usuario_atual,
                    setor

                )

                VALUES (

                    :id_maquina,
                    :marca,
                    :modelo,
                    :sistema,
                    :ram,
                    :armazenamento,
                    :usuario,
                    :setor

                )

            """),

            {

                "id_maquina":
                request.form.get("id_maquina"),

                "marca":
                request.form.get("marca"),

                "modelo":
                request.form.get("modelo"),

                "sistema":
                request.form.get(
                    "sistema_operacional"
                ),

                "ram":
                request.form.get("memoria_ram"),

                "armazenamento":
                request.form.get(
                    "armazenamento"
                ),

                "usuario":
                request.form.get(
                    "usuario_atual"
                ),

                "setor":
                request.form.get("setor")

            }

        )

        db.session.commit()

        flash(
            "Ativo criado!",
            "success"
        )

        return redirect("/inventario")

    return render_template(
        "novo_ativo.html",
        usuario=usuario
    )
# =========================
# EXCLUIR ATIVO
# =========================

@app.route(
    "/excluir_ativo/<int:id>"
)
def excluir_ativo(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    db.session.execute(

        db.text("""

            DELETE FROM ativos

            WHERE id = :id

        """),

        {
            "id": id
        }

    )

    db.session.commit()

    flash(
        "Ativo excluído!",
        "success"
    )

    return redirect("/inventario")

# =========================
# ESTOQUE TI
# =========================

@app.route("/estoque")
def estoque():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    itens = db.session.execute(

        db.text("""

            SELECT *

            FROM estoque_ti

            ORDER BY categoria, item

        """)

    ).fetchall()

    movimentacoes = db.session.execute(

        db.text("""

            SELECT

                m.id,
                m.tipo,
                m.quantidade,
                m.usuario,
                m.observacao,
                m.criado_em,
                e.item,
                e.categoria

            FROM estoque_movimentacoes m

            JOIN estoque_ti e
            ON e.id = m.item_id

            ORDER BY m.criado_em DESC

            LIMIT 20

        """)

    ).fetchall()

    total_itens = db.session.execute(

        db.text("""

            SELECT COUNT(*)

            FROM estoque_ti

        """)

    ).scalar()

    itens_baixo = db.session.execute(

        db.text("""

            SELECT COUNT(*)

            FROM estoque_ti

            WHERE quantidade <= minimo

        """)

    ).scalar()

    total_unidades = db.session.execute(

        db.text("""

            SELECT COALESCE(SUM(quantidade), 0)

            FROM estoque_ti

        """)

    ).scalar()

    return render_template(

        "estoque.html",

        usuario=usuario,

        itens=itens,

        movimentacoes=movimentacoes,

        total_itens=total_itens,

        itens_baixo=itens_baixo,

        total_unidades=total_unidades

    )


# =========================
# NOVO ITEM ESTOQUE
# =========================

@app.route(
    "/novo_item_estoque",
    methods=["POST"]
)
def novo_item_estoque():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    db.session.execute(

        db.text("""

            INSERT INTO estoque_ti (

                item,
                categoria,
                quantidade,
                minimo,
                observacao

            )

            VALUES (

                :item,
                :categoria,
                :quantidade,
                :minimo,
                :observacao

            )

        """),

        {
            "item":
            request.form.get("item", ""),

            "categoria":
            request.form.get("categoria", ""),

            "quantidade":
            int(request.form.get("quantidade") or 0),

            "minimo":
            int(request.form.get("minimo") or 1),

            "observacao":
            request.form.get("observacao", "")
        }

    )

    db.session.commit()

    flash(
        "Item adicionado ao estoque!",
        "success"
    )

    return redirect("/estoque")

# =========================
# EDITAR ITEM ESTOQUE
# =========================

@app.route(
    "/editar_item_estoque/<int:id>",
    methods=["GET", "POST"]
)
def editar_item_estoque(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    item = db.session.execute(

        db.text("""

            SELECT *

            FROM estoque_ti

            WHERE id = :id

        """),

        {
            "id": id
        }

    ).fetchone()

    if not item:

        flash(
            "Item não encontrado.",
            "error"
        )

        return redirect("/estoque")

    if request.method == "POST":

        db.session.execute(

            db.text("""

                UPDATE estoque_ti

                SET

                    item = :item,
                    categoria = :categoria,
                    quantidade = :quantidade,
                    minimo = :minimo,
                    observacao = :observacao

                WHERE id = :id

            """),

            {
                "id": id,

                "item":
                request.form.get("item", ""),

                "categoria":
                request.form.get("categoria", ""),

                "quantidade":
                int(request.form.get("quantidade") or 0),

                "minimo":
                int(request.form.get("minimo") or 1),

                "observacao":
                request.form.get("observacao", "")
            }

        )

        db.session.commit()

        flash(
            "Item atualizado com sucesso!",
            "success"
        )

        return redirect("/estoque")

    return render_template(

        "editar_item_estoque.html",

        usuario=usuario,

        item=item

    )


# =========================
# MOVIMENTAR ESTOQUE
# =========================

@app.route(
    "/movimentar_estoque",
    methods=["POST"]
)
def movimentar_estoque():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    item_id = request.form.get("item_id")
    tipo = request.form.get("tipo")
    quantidade = int(request.form.get("quantidade") or 0)
    pessoa = request.form.get("usuario", "")
    observacao = request.form.get("observacao", "")

    if quantidade <= 0:

        flash(
            "Quantidade inválida.",
            "error"
        )

        return redirect("/estoque")

    item = db.session.execute(

        db.text("""

            SELECT *

            FROM estoque_ti

            WHERE id = :id

        """),

        {
            "id": item_id
        }

    ).fetchone()

    if not item:

        flash(
            "Item não encontrado.",
            "error"
        )

        return redirect("/estoque")

    if tipo == "saida" and item.quantidade < quantidade:

        flash(
            "Estoque insuficiente para esta saída.",
            "error"
        )

        return redirect("/estoque")

    if tipo == "entrada":

        nova_quantidade = item.quantidade + quantidade

    elif tipo == "saida":

        nova_quantidade = item.quantidade - quantidade

    else:

        flash(
            "Tipo de movimentação inválido.",
            "error"
        )

        return redirect("/estoque")

    db.session.execute(

        db.text("""

            UPDATE estoque_ti

            SET quantidade = :quantidade

            WHERE id = :id

        """),

        {
            "quantidade": nova_quantidade,
            "id": item_id
        }

    )

    db.session.execute(

        db.text("""

            INSERT INTO estoque_movimentacoes (

                item_id,
                tipo,
                quantidade,
                usuario,
                observacao

            )

            VALUES (

                :item_id,
                :tipo,
                :quantidade,
                :usuario,
                :observacao

            )

        """),

        {
            "item_id": item_id,
            "tipo": tipo,
            "quantidade": quantidade,
            "usuario": pessoa,
            "observacao": observacao
        }

    )

    db.session.commit()

    flash(
        "Movimentação registrada!",
        "success"
    )

    return redirect("/estoque")

# =========================
# DESFAZER MOVIMENTAÇÃO ESTOQUE
# =========================

@app.route("/desfazer_movimentacao_estoque/<int:id>")
def desfazer_movimentacao_estoque(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    mov = db.session.execute(

        db.text("""

            SELECT *

            FROM estoque_movimentacoes

            WHERE id = :id

        """),

        {
            "id": id
        }

    ).fetchone()

    if not mov:

        flash(
            "Movimentação não encontrada.",
            "error"
        )

        return redirect("/estoque")

    item = db.session.execute(

        db.text("""

            SELECT *

            FROM estoque_ti

            WHERE id = :id

        """),

        {
            "id": mov.item_id
        }

    ).fetchone()

    if not item:

        flash(
            "Item relacionado não encontrado.",
            "error"
        )

        return redirect("/estoque")

    if mov.tipo == "saida":

        nova_quantidade = item.quantidade + mov.quantidade

    elif mov.tipo == "entrada":

        nova_quantidade = item.quantidade - mov.quantidade

        if nova_quantidade < 0:

            flash(
                "Não foi possível desfazer: quantidade ficaria negativa.",
                "error"
            )

            return redirect("/estoque")

    else:

        flash(
            "Tipo de movimentação inválido.",
            "error"
        )

        return redirect("/estoque")

    db.session.execute(

        db.text("""

            UPDATE estoque_ti

            SET quantidade = :quantidade

            WHERE id = :id

        """),

        {
            "quantidade": nova_quantidade,
            "id": mov.item_id
        }

    )

    db.session.execute(

        db.text("""

            DELETE FROM estoque_movimentacoes

            WHERE id = :id

        """),

        {
            "id": id
        }

    )

    db.session.commit()

    flash(
        "Movimentação desfeita com sucesso!",
        "success"
    )

    return redirect("/estoque")


# =========================
# EXCLUIR ITEM ESTOQUE
# =========================

@app.route("/excluir_item_estoque/<int:id>")
def excluir_item_estoque(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    db.session.execute(

        db.text("""

            DELETE FROM estoque_movimentacoes

            WHERE item_id = :id

        """),

        {
            "id": id
        }

    )

    db.session.execute(

        db.text("""

            DELETE FROM estoque_ti

            WHERE id = :id

        """),

        {
            "id": id
        }

    )

    db.session.commit()

    flash(
        "Item removido do estoque.",
        "success"
    )

    return redirect("/estoque")

# =========================
# HISTÓRICO MAPA
# =========================

@app.route("/historico_mapa")
def historico_mapa():

    usuario = session.get("usuario")

    if not usuario:

        return redirect("/login")

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:

        return redirect("/")

    historico = db.session.execute(

        db.text("""

            SELECT *

            FROM movimentacoes_mapa

            ORDER BY data_movimentacao DESC

        """)

    ).fetchall()

    return render_template(

        "historico_mapa.html",

        usuario=usuario,

        historico=historico

    )

# =========================
# HISTÓRICO DA MÁQUINA
# =========================

@app.route("/historico_maquina/<id_maquina>")
def historico_maquina(id_maquina):

    usuario = session.get("usuario")

    if not usuario:

        return redirect("/login")

    ativo = db.session.execute(

        db.text("""

            SELECT *

            FROM ativos

            WHERE id_maquina = :id_maquina

        """),

        {
            "id_maquina": id_maquina
        }

    ).fetchone()

    if not ativo:

        flash(
            "Máquina não encontrada.",
            "error"
        )

        return redirect("/")

    chamados = db.session.execute(

        db.text("""

            SELECT *

            FROM chamados

            WHERE maquina = :id_maquina

            ORDER BY id DESC

        """),

        {
            "id_maquina": id_maquina
        }

    ).fetchall()

    movimentacoes = db.session.execute(

        db.text("""

            SELECT *

            FROM movimentacoes_mapa

            WHERE maquina = :id_maquina

            ORDER BY data_movimentacao DESC

        """),

        {
            "id_maquina": id_maquina
        }

    ).fetchall()

    total_chamados = len(chamados)

    chamados_abertos = len([

        c for c in chamados

        if (c.status or "").lower() != "finalizado"

    ])

    chamados_finalizados = len([

        c for c in chamados

        if (c.status or "").lower() == "finalizado"

    ])

    return render_template(

        "historico_maquina.html",

        usuario=usuario,

        ativo=ativo,

        chamados=chamados,

        movimentacoes=movimentacoes,

        total_chamados=total_chamados,

        chamados_abertos=chamados_abertos,

        chamados_finalizados=chamados_finalizados

    )
# =========================
# KANBAN - TAREFAS TI
# =========================

# =========================
# TAREFAS
# =========================

@app.route("/tarefas")
def tarefas():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    pendentes = db.session.execute(

        db.text("""

            SELECT *

            FROM tarefas

            WHERE LOWER(status) = 'pendente'

            ORDER BY id DESC

        """)

    ).fetchall()

    andamento = db.session.execute(

        db.text("""

            SELECT *

            FROM tarefas

            WHERE LOWER(status) = 'andamento'

            ORDER BY id DESC

        """)

    ).fetchall()

    aguardando = db.session.execute(

        db.text("""

            SELECT *

            FROM tarefas

            WHERE LOWER(status) = 'aguardando'

            ORDER BY id DESC

        """)

    ).fetchall()

    finalizadas = db.session.execute(

        db.text("""

            SELECT *

            FROM tarefas

            WHERE LOWER(status) = 'finalizado'

            ORDER BY id DESC

        """)

    ).fetchall()

    return render_template(

        "tarefas.html",

        usuario=usuario,

        pendentes=pendentes,

        andamento=andamento,

        aguardando=aguardando,

        finalizadas=finalizadas

    )


# =========================
# NOVA TAREFA
# =========================

@app.route(
    "/nova_tarefa",
    methods=["POST"]
)
def nova_tarefa():

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    dados = {

        "titulo":
        request.form.get("titulo"),

        "descricao":
        request.form.get("descricao"),

        "prioridade":
        request.form.get("prioridade"),

        "responsavel":
        request.form.get("responsavel"),

        "setor":
        request.form.get("setor"),

        "prazo":
        request.form.get("prazo") or None,

        "criado_por":
        usuario["nome"]

    }

    db.session.execute(

        db.text("""

            INSERT INTO tarefas (

                titulo,
                descricao,
                prioridade,
                responsavel,
                setor,
                prazo,
                status,
                criado_por

            )

            VALUES (

                :titulo,
                :descricao,
                :prioridade,
                :responsavel,
                :setor,
                :prazo,
                'pendente',
                :criado_por

            )

        """),

        dados

    )

    # =========================
    # NOTIFICAÇÃO
    # =========================

    responsavel_nome = request.form.get("responsavel")

    if responsavel_nome:

        responsavel_db = db.session.execute(

            db.text("""

                SELECT id

                FROM usuarios

                WHERE LOWER(nome) = LOWER(:nome)

                LIMIT 1

            """),

            {
                "nome": responsavel_nome
            }

        ).fetchone()

        if responsavel_db:

            criar_notificacao(

                usuario_id=responsavel_db.id,

                titulo="Nova tarefa atribuída",

                mensagem=f"Você recebeu a tarefa: {request.form.get('titulo')}",

                tipo="kanban",

                link="/tarefas"

            )

    db.session.commit()

    flash(
        "Tarefa criada com sucesso!",
        "success"
    )

    return redirect("/tarefas")
# =========================
# ALTERAR STATUS
# =========================

@app.route("/mover_tarefa/<int:id>/<status>")
def mover_tarefa(id, status):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    status_validos = [
        "pendente",
        "andamento",
        "aguardando",
        "finalizado"
    ]

    if status not in status_validos:

        flash(
            "Status inválido.",
            "error"
        )

        return redirect("/tarefas")

    db.session.execute(

        db.text("""

            UPDATE tarefas

            SET status = :status

            WHERE id = :id

        """),

        {
            "id": id,
            "status": status
        }

    )

    db.session.commit()

    flash(
        "Status atualizado!",
        "success"
    )

    return redirect("/tarefas")


# =========================
# EDITAR TAREFA
# =========================

@app.route(
    "/editar_tarefa/<int:id>",
    methods=["GET", "POST"]
)
def editar_tarefa(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    tarefa = db.session.execute(

        db.text("""

            SELECT *

            FROM tarefas

            WHERE id = :id

        """),

        {
            "id": id
        }

    ).fetchone()

    if not tarefa:

        flash(
            "Tarefa não encontrada.",
            "error"
        )

        return redirect("/tarefas")

    if request.method == "POST":

        dados = {

            "id": id,

            "titulo":
            request.form.get("titulo"),

            "descricao":
            request.form.get("descricao"),

            "prioridade":
            request.form.get("prioridade"),

            "responsavel":
            request.form.get("responsavel"),

            "setor":
            request.form.get("setor"),

            "prazo":
            request.form.get("prazo") or None

        }

        db.session.execute(

            db.text("""

                UPDATE tarefas

                SET

                    titulo = :titulo,
                    descricao = :descricao,
                    prioridade = :prioridade,
                    responsavel = :responsavel,
                    setor = :setor,
                    prazo = :prazo

                WHERE id = :id

            """),

            dados

        )

        responsavel_db = None

        if dados["responsavel"]:

            responsavel_db = db.session.execute(

                db.text("""

                    SELECT id

                    FROM usuarios

                    WHERE LOWER(nome) = LOWER(:nome)

                    LIMIT 1

                """),

                {
                    "nome": dados["responsavel"]
                }

            ).fetchone()

        if responsavel_db:

            criar_notificacao(

                usuario_id=responsavel_db.id,

                titulo="Tarefa atualizada",

                mensagem=f"A tarefa '{dados['titulo']}' foi atualizada por {usuario['nome']}.",

                tipo="kanban",

                link="/tarefas"

            )

        db.session.commit()

        flash(
            "Tarefa atualizada!",
            "success"
        )

        return redirect("/tarefas")

    return render_template(

        "editar_tarefa.html",

        usuario=usuario,

        tarefa=tarefa

    )

# =========================
# EXCLUIR TAREFA
# =========================

@app.route("/excluir_tarefa/<int:id>")
def excluir_tarefa(id):

    usuario = session.get("usuario")

    if not usuario:
        return redirect("/login")

    if usuario["tipo"] != "ti":
        return redirect("/")

    tarefa = db.session.execute(

        db.text("""

            SELECT *

            FROM tarefas

            WHERE id = :id

        """),

        {
            "id": id
        }

    ).fetchone()

    if not tarefa:

        flash(
            "Tarefa não encontrada.",
            "error"
        )

        return redirect("/tarefas")

    responsavel_db = None

    if tarefa.responsavel:

        responsavel_db = db.session.execute(

            db.text("""

                SELECT id

                FROM usuarios

                WHERE LOWER(nome) = LOWER(:nome)

                LIMIT 1

            """),

            {
                "nome": tarefa.responsavel
            }

        ).fetchone()

    if responsavel_db:

        criar_notificacao(

            usuario_id=responsavel_db.id,

            titulo="Tarefa excluída",

            mensagem=f"A tarefa '{tarefa.titulo}' foi excluída por {usuario['nome']}.",

            tipo="kanban",

            link="/tarefas"

        )

    db.session.execute(

        db.text("""

            DELETE FROM tarefas

            WHERE id = :id

        """),

        {
            "id": id
        }

    )

    db.session.commit()

    flash(
        "Tarefa excluída!",
        "success"
    )

    return redirect("/tarefas")

# =========================
# TESTE BANCO
# =========================

@app.route("/teste_bd")
def teste_bd():

    try:

        db.session.execute(
        db.text("SELECT 1")
    )

        return "Banco conectado com sucesso!"

    except Exception as e:

        return f"Erro: {e}"
@app.route("/debug-db")
def debug_db():
    result = db.session.execute(
        db.text("SELECT COUNT(*) FROM usuarios")
    ).fetchone()
    return f"Usuarios: {result[0]}"

@app.route("/reset_admin")
def reset_admin():

    senha_hash = generate_password_hash("123")

    db.session.execute(db.text("""

        UPDATE usuarios
        SET senha = :senha
        WHERE nome = 'admin'

    """), {
        "senha": senha_hash
    })

    db.session.commit()

    return "Senha resetada!"

# =========================
# API INVENTÁRIO
# =========================
@app.route("/api/inventario")
def api_inventario():

    usuario = session.get("usuario")

    if not usuario:

        return {
            "erro": "Não autenticado"
        }, 401

    if usuario["tipo"] not in [
        "ti",
        "administracao"
    ]:

        return {
            "erro": "Sem permissão"
        }, 403

    dados = db.session.execute(

        db.text("""

            SELECT

                usuario_atual,
                id_maquina,
                marca,
                modelo,
                sistema_operacional,
                memoria_ram,
                armazenamento,
                setor,
                status,
                observacoes

            FROM ativos

            ORDER BY usuario_atual

        """)

    ).fetchall()

    inventario = {}

    for item in dados:

        chave = item.usuario_atual

        if not chave:

            chave = "SEM_USUARIO"

        if chave not in inventario:

            inventario[chave] = []

        inventario[chave].append({

            "patrimonio": item.id_maquina,
            "marca": item.marca,
            "modelo": item.modelo,
            "sistema_operacional": item.sistema_operacional,
            "memoria_ram": item.memoria_ram,
            "armazenamento": item.armazenamento,
            "setor": item.setor,
            "status": item.status,
            "observacoes": item.observacoes

        })

    return inventario

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )

